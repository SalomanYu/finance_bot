import requests
import re
import json
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
import os
import openpyxl
from datetime import datetime

from base import create_table

def get_courses_data(text):
    # sourse = open('page_sourse.html', 'r').read()

    # index = requests.get('https://skillfactory.ru/catalogue').text
    id_serach = re.findall("storepart: '\w+'", text)

    # data = []
    for id in id_serach:
        get_url = f'https://store.tildacdn.com/api/getproductslist/?storepartuid={id[12:24]}'
        json_result = requests.get(get_url).text
        save_data(json_result, id)

        
def save_data(data, filename):
    file = open(f'JSON/{filename}_data.json', 'w')
    json.dump(json.loads("".join(data)), file, ensure_ascii=False, indent=2)
    print('Информация о курсах загружена!')

def save_page():
    url = 'https://skillfactory.ru/catalogue'
    driver = webdriver.Chrome()
    driver.get(url)
    driver.implicitly_wait(10)

    more_courses_btn = driver.find_element(By.XPATH, "//div[@class='js-store-load-more-btn t-store__load-more-btn t-btn t-btn_sm']")
    more_courses_btn.click()
    sleep(2)
    
    result = driver.page_source
    get_courses_data(result)


class SkillFactory:
    def __init__(self, outputfilename):
        self.outputfilename = outputfilename
        self.row_count = 2

    def start(self):
        create_table(self.outputfilename)
        print('Успешно! Таблица создана')
        self.workbook_writer = openpyxl.load_workbook(self.outputfilename)
        self.sheet_writer = self.workbook_writer.worksheets[0]

        self.__parse_json()

    def __parse_json(self):
        for file in os.listdir('JSON'):
            json_data = json.load(open(f'JSON/{file}'))
            products = json_data['products']
            for course in products:
                self.title = course['title']
                self.url = course['url']
                self.description = self.edit_description(course['descr'])
                self.price = float(course['price']) if course['price'] not in ('', None) else ''
                self.oldprice = float(course['priceold']) if course['priceold'] != '' else ''
                self.image = self.edit_image(course['gallery'])
                
                for item in course['characteristics']:
                    if item['title'] == 'Направление':
                        self.subcategory = item['value']
                        break
                self.add_row_to_table()
                print(self.title)
        self.workbook_writer.save(self.outputfilename)

    def edit_image(self, img):
        result = json.loads((img.strip('][').replace('\\', '')))
        return result['img']
    
    def edit_description(self, descr):
        result = descr.split('<br /><br />')[0]
        return result
    
    def add_row_to_table(self):
        self.sheet_writer.cell(self.row_count, 1).value = 'SkillFactory'
        self.sheet_writer.cell(self.row_count, 2).value = self.title.replace('»', '').replace('«','')
        self.sheet_writer.cell(self.row_count, 5).value = self.description
        self.sheet_writer.cell(self.row_count, 8).value = self.subcategory
        self.sheet_writer.cell(self.row_count, 16).value = 'Русский'
        self.sheet_writer.cell(self.row_count, 18).value = self.url        
        self.sheet_writer.cell(self.row_count, 31).value = self.image
        self.sheet_writer.cell(self.row_count, 33).value = self.oldprice
        self.sheet_writer.cell(self.row_count, 34).value = self.price

        self.row_count += 1

date = datetime.now()    
bot = SkillFactory(f'SkillFactory({date.day}.{date.month}.{date.year}).xlsx')
bot.start()
# bot.edit_image('[{\"img\":\"https:\\/\\/static.tildacdn.com\\/tild3434-3363-4939-b862-316532323031\\/Bages_1_1200x630_-_1.png\"}]')