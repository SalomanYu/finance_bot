from bs4 import BeautifulSoup, Tag
from numpy import NaN
import requests
import pandas as pd
import re
import xlsxwriter
import openpyxl
import math

class Course:
    def __init__(self, exfilename, outputname='EXCEL/result'):
        self.excel_path = exfilename
        self.outputname = outputname
        self.count = 2
        self.headers = {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.99 Safari/537.36'}
    
    def run(self):
        self.create_table()
        self.connect_to_excel()
        # for item in range(len(self.all_urls)-1):
        for item in range(1, len(self.all_urls)):
            self.url = self.all_urls[item]
            if 'https:' in str(self.url):
                self.course_name = self.all_names[item]
                self.course_desc = self.all_descriptions[item]
                self.course_category = self.all_categoryes[item]
                self.course_picture = self.all_pictures[item]
                self.course_price = self.all_prices[item]
                try:
                    print(self.url, 'ПРОШЕЛ')
                    req = requests.get(self.url, headers=self.headers)
                    self.soup = BeautifulSoup(req.text, 'lxml')
                    self.add_row()
                except (requests.exceptions.MissingSchema) as err:
                    print(err)
            else:
                print('Не прошел', self.url)


    def connect_to_excel(self):
        workbook = pd.ExcelFile(self.excel_path)
        sheet = workbook.parse(0)

        self.all_urls = sheet['url'][1:]
        self.all_names = sheet['name'][1:]
        self.all_prices = sheet['price'][1:]
        self.all_pictures = sheet['picture'][1:]
        self.all_categoryes = sheet['categoryId'][1:]
        self.all_descriptions = sheet['description'][1:]

    def add_row(self):
        workbook = openpyxl.load_workbook(f'{self.outputname}.xlsx')
        sheet = workbook.worksheets[0]
        sheet.cell(self.count, 1).value = 'Uchmet'
        sheet.cell(self.count, 2).value =  self.course_name
        sheet.cell(self.count, 5).value = self.course_desc
        sheet.cell(self.count, 10).value = self.get_format()
        sheet.cell(self.count, 11).value = self.course_category
        sheet.cell(self.count, 12).value = 'Русский'
        sheet.cell(self.count, 14).value = self.url
        sheet.cell(self.count, 18).value = 'Есть'
        sheet.cell(self.count, 27).value = self.course_picture
        sheet.cell(self.count, 28).value = ''
        sheet.cell(self.count, 30).value = self.course_price
        sheet.cell(self.count, 35).value = " ".join(self.get_program())
        sheet.cell(self.count, 36).value = " | ".join(self.get_teachers()[2])
        sheet.cell(self.count, 37).value = " | ".join(self.get_teachers()[0])
        sheet.cell(self.count, 38).value = " | ".join(self.get_teachers()[1])
        print(f"[{self.count-1}] {self.course_name}")
        self.count += 1
        workbook.save(f'{self.outputname}.xlsx')

    def get_program(self):
        try:
            count_ul = self.soup.find('div', class_='page-block__content measure-detail__block').find_all('ul')[1]
            course_program = []
            for item in count_ul:
                if isinstance(item, Tag):
                    module_all_text = item.text.strip().split('\n')
                    module_title = "<h3>" + module_all_text[0] + "</h3>"
                    module_content = ["<li>" + item + "</li>" for item in module_all_text[1:]]
                    try:
                        course_program.append(' '.join(module_title, " ".join(module_content)))
                    except TypeError:
                        course_program.append(module_title)
                
            return course_program      

        except (IndexError, AttributeError):
            try:
                program_content = self.soup.find('div', class_='page-block__content measure-detail__block')                     
                program_elems = re.findall(r'\d\d\. .*?\n|\d\. .*?\n', program_content.text) # Находим строки начинающиеся на 1. Название
                result = ["<p>" + item + '</p>' for item in program_elems]
            except:
                return []
            if result == []:
                try:
                    program = ["<li>"+item.text.strip()+"</li>" for item in program_content.find('ul').find_all('li')]
                    return program
                except AttributeError:
                    try:
                        program = ["<li>"+item.text.strip()+"</li>" for item in program_content.find('ol').find_all('li')]
                    except AttributeError:
                        return []
                    return program
            else:
                return result

    def get_format(self):
        try:
            format = self.soup.find_all('div', class_='el-block')[2].find('a').text.strip()
        except (AttributeError, IndexError):
            try:
                format = self.soup.find_all('div', class_='el-block')[3].find('a').text.strip()
            except IndexError:
                return 'офлайн'
        return format
    
    def for_whom(self):
        try:
            who_block = self.soup.find('i', class_='fa fa-user fa-fw').parent
            who = [item.text for item in who_block.find_all('span')]
        except AttributeError:
            return ""
        return " ".join(who)

    def get_teachers(self):
        try:
            urls = self.soup.find_all('div', class_='el-block')[-1].find_all('div', class_='el')[-1].find_all('a')
        except IndexError:
            return [], [], []
        names = []
        descriptions = []
        avatars = []
        for item in urls:
            url_teacher_info = "https://www.uchmet.ru" + item['href']
            req = requests.get(url_teacher_info)
            soup = BeautifulSoup(req.text, 'lxml')
            try: 
                avatar = "https://www.uchmet.ru" + soup.find('img', class_='img-responsive center-block img-autor')['src']
                description = [item.text for item in soup.find('div', class_='col-md-9').find_all('p')]
                teacher_name = item.text
                names.append(teacher_name)
                descriptions.append(" ".join(description))
                avatars.append(avatar)
            
            except TypeError:
                description = ''
                avatar = ''
                all_paragraphs = self.soup.find('div', class_='page-block__content measure-detail__block').find_all('p')
                for item in all_paragraphs:
                    if 'Ведущий вебинара' in str(item):
                        try:
                            teacher_name = item.i.text
                        except AttributeError:
                            try:
                                teacher_name = item.em.text
                            except AttributeError:
                                teacher_name = ''
                        try:
                            return [teacher_name], [description], [avatar]
                        except:
                            print(self.url)
                            quit()

            
        return names, descriptions, avatars

    def create_table(self):
        workbook = xlsxwriter.Workbook(f"{self.outputname}.xlsx")
        sheet = workbook.add_worksheet()

        sheet.write(0, 0, 'Организатор курса')
        sheet.write(0, 1, 'Название курса')
        sheet.write(0, 2, 'Популярные')
        sheet.write(0, 3, 'Рекомендуемые')
        sheet.write(0, 4, 'Описание курса')
        sheet.write(0, 5, 'Выбор подкатегории')
        sheet.write(0, 6, 'Уровень')
        sheet.write(0, 7, 'Баллы')
        sheet.write(0, 8, 'Сертификат')
        sheet.write(0, 9, 'Формат')
        sheet.write(0, 10, 'Тип')
        sheet.write(0, 11, 'Язык')
        sheet.write(0, 12, 'Продолжительность в неделях')
        sheet.write(0, 13, 'Ссылка на страницу')
        sheet.write(0, 14, 'Вариант для пропуска')
        sheet.write(0, 15, 'Дата начала')
        sheet.write(0, 16, 'Удостоверение о повышении клалификации')
        sheet.write(0, 17, 'Сертификат о прохождении курса')
        sheet.write(0, 18, 'Диплом об окончании')
        sheet.write(0, 19, 'Практико-ориентированность')
        sheet.write(0, 20, 'Наставник')
        sheet.write(0, 21, 'Помощь в трудоустройстве')
        sheet.write(0, 22, 'Наработка портфолио')
        sheet.write(0, 23, 'Рассрочка')
        sheet.write(0, 24, 'Налоговый вычет')
        sheet.write(0, 25, 'Обложка курса')
        sheet.write(0, 26, 'Картинка курса')
        sheet.write(0, 27, 'Приветственное видео')
        sheet.write(0, 28, 'Старая цена')
        sheet.write(0, 29, 'Цена')
        sheet.write(0, 30, 'Срок рассрочки в месяцах')
        sheet.write(0, 31, 'Платеж по рассрочке в рублях')
        sheet.write(0, 32, 'Навыки')
        sheet.write(0, 33, 'Программное обеспечение')
        sheet.write(0, 34, 'Программа курса')
        sheet.write(0, 35, 'Фото преподавателя')
        sheet.write(0, 36, 'ФИО преподавателя')
        sheet.write(0, 37, 'О преподавателе')
        sheet.write(0, 38, 'Описание преподавателя')
        sheet.write(0, 39, 'Текстовое поле ввода данных1')
        sheet.write(0, 40, 'Текстовое поле ввода данных1_1')
        sheet.write(0, 41, 'Загрузка картинки jpeg')
        sheet.write(0, 42, 'Кол-во уроков')
        sheet.write(0, 43, 'Предмет')
        sheet.write(0, 44, 'Требуемые знания')
        sheet.write(0, 45, 'Партнеры')

        workbook.close()

from datetime import datetime

date = datetime.now()
bot = Course(exfilename='uchmet_parthers_may.xls', outputname=f'Uchmet({date.day}.{date.month}.{date.year})')
bot.run()
# print(bot.get_program())