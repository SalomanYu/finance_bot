from base import create_table

import requests
from bs4 import BeautifulSoup, Tag
import json
import re
from math import ceil
import openpyxl

class XYZ:
    def __init__(self, outputfilename):
        self.outputfilename = outputfilename
        self.row_count = 2
    

    def start(self):
        create_table(self.outputfilename)
        print('Успешно! Таблица создана')
        self.workbook_writer = openpyxl.load_workbook(self.outputfilename)
        self.sheet_writer = self.workbook_writer.worksheets[0]

        request = requests.get('https://www.school-xyz.com/courses#!all/all')
        self.soup = BeautifulSoup(request.text, 'lxml')

        self.parse_all_courses()

        self.workbook_writer.save(self.outputfilename)

    def collect_prices_by_api(self):
        req = requests.get('https://learn.school-xyz.com/api/public/v1/courses/')
        all_courses_price = json.loads(req.text)

        return all_courses_price    

    def parse_all_courses(self):
        pattern = 'course-grid__item js-course-category.*?\\n'
        all_courses_price = self.collect_prices_by_api() 
        all_courses = list(set(re.findall(pattern, str(self.soup.find('div', class_='page-courses__grid course-grid js-courses')))))
        for class_name in all_courses:
            category_course = self.soup.find_all('div', class_=class_name.replace('">\n', ''))
            for course in category_course:
                if isinstance(course, Tag):
                    self.title = course.find('h3', class_='card-course__title').text
                    self.url = course.a['href']
                    self.img = course.img['src']
                    self.description = course.find('div', class_='card-course__description').text
                    self.category = course.find('div', class_='card-course__category').text
                    
                    try:
                        self.start_date = course.find('div', class_='xyz-course-nextClassStartDate').text.replace('Начало ', '')
                    except:
                        self.start_date = ''

                    self.course_info = course.find_all('div', class_='card-course__info-item')
                    if len(self.course_info) > 1:
                        self.direction = 4 * ceil(float(self.course_info[0].text.replace(',', '.').split(' ')[0]))
                    else:
                        self.direction = ''
                    
                    self.price = ''
                    self.oldprice = ''

                    for api_course in all_courses_price:
                        if api_course['title'] == self.title:
                            self.price = api_course['regions']['ru']['price']
                            self.oldprice = api_course['regions']['ru']['oldPrice']


                    self.add_row_to_table()

    def select_subcategory(self, category):
        categories_dict = [
            {
                'keys': ('3D-моделирование', '2D/Концепт-арт', 'Геймдизайн'),
                'values':{ 
                    'category': 'Дизайн',
                    'subcategory': 'Графический дизайн|Дизайн игр'
                    }
            },
            {
                'keys': ('Программирование',),
                'values':{ 
                    'category': 'IT',
                    'subcategory': 'Разработка игр'
                    },
            },
                {
                'keys': ('VFX',),
                'values':{ 
                    'category': 'Дизайн',
                    'subcategory': 'Другое'
                    },
            },
                {
                'keys': ('Бизнес',),
                'values':{ 
                    'category': 'Бизнес',
                    'subcategory': 'Менеджмент'
                }
            }
        ]
        # print(len(categories_dict))
        for group in categories_dict:
            if category in group['keys']:
                res_category = group['values']['category']
                res_subcategory = group['values']['subcategory']
                return res_category, res_subcategory

        return '', ''

    def add_row_to_table(self):
        self.sheet_writer.cell(self.row_count, 1).value = 'XYZ SCHOOL'
        self.sheet_writer.cell(self.row_count, 2).value = self.title
        self.sheet_writer.cell(self.row_count, 5).value = self.description
        self.sheet_writer.cell(self.row_count, 6).value = self.select_subcategory(self.category)[0]
        self.sheet_writer.cell(self.row_count, 8).value = self.select_subcategory(self.category)[1]
        self.sheet_writer.cell(self.row_count, 16).value = 'Русский'
        self.sheet_writer.cell(self.row_count, 18).value = self.url

        self.sheet_writer.cell(self.row_count, 31).value = self.img
        self.sheet_writer.cell(self.row_count, 33).value = self.oldprice
        self.sheet_writer.cell(self.row_count, 34).value = self.price

        self.row_count += 1   


bot = XYZ('XYZ.xlsx')
bot.start()