import re
from random import choice
import openpyxl
import requests 
from bs4 import BeautifulSoup
import json
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
from datetime import datetime

from base import create_table

# https://practicum.yandex.ru/api/v2/professions/
# Ссылка с которой был стянут Json

class YandexPractium:
    def __init__(self, outputfilename):
        self.outputfilename = outputfilename
        self.domain = 'https://practicum.yandex.ru/'
        self.api_url = 'https://practicum.yandex.ru/api/v2/professions/'
        self.row_count = 2        
        # self.headers = {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.99 Safari/537.36'}
        self.headers = {'Googlebot': 'Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)'}

    def start(self):
        # self.course_url = 'https://practicum.yandex.ru/sql-data-analyst/'
        # req = requests.get(self.course_url, headers=self.headers)
        # self.soup = BeautifulSoup(req.text, 'lxml')

        create_table(self.outputfilename)
        print('Успешно! Таблица создана')
        self.workbook_writer = openpyxl.load_workbook(self.outputfilename)
        self.sheet_writer = self.workbook_writer.worksheets[0]


        all_courses_info = json.load(open('all_courses_test.json', 'r'))
        for course in all_courses_info:
            self.course_name = course['name']
            self.duration = course['duration']
            self.price = course['price'] if course['price'] != 0 else ''
            
            try:
                self.subcategory = course['tags'][0]['name']
            except IndexError:
                self.subcategory = ''

            self.course_url = course['landing_path']
            if self.course_url == '':
                self.course_url = self.domain + course['slug']
            elif 'yandex.ru/' not in self.course_url:
                self.course_url = self.domain + course['landing_path'].replace('/', '')

            # print(choice(self.headers))
            request = requests.get(self.course_url, headers=self.headers)
            self.soup = BeautifulSoup(request.text, 'lxml')

            self.description = self.get_description()
            if self.description == '':
                self.description = course['description'] if course['description'] != '-' else ''
            elif self.description == 'Empty description':
                self.description = ''
            self.add_row_to_table()
        
        self.workbook_writer.save(self.outputfilename)


    def add_row_to_table(self):

        self.sheet_writer.cell(self.row_count, 1).value = 'Яндекс.Практикум'
        self.sheet_writer.cell(self.row_count, 2).value = self.course_name
        self.sheet_writer.cell(self.row_count, 5).value = self.description
        self.sheet_writer.cell(self.row_count, 8).value = self.subcategory
        self.sheet_writer.cell(self.row_count, 16).value = 'Русский'
        self.sheet_writer.cell(self.row_count, 18).value = self.course_url

        self.sheet_writer.cell(self.row_count, 31).value = self.get_image()
        self.sheet_writer.cell(self.row_count, 34).value = self.price

        program = json.dumps(self.get_program(), ensure_ascii=False)
        
        print(len(program), self.course_name)
        self.sheet_writer.cell(self.row_count, 40).value = program
        self.sheet_writer.cell(self.row_count, 39).value = self.is_module

        self.sheet_writer.cell(self.row_count, 37).value = self.get_skills()

        self.row_count += 1   

    def get_description(self):
        try:
            descr = self.soup.find('div', class_='landing-grid__column landing-grid__column_size_full about-profession-section__description').text
        except AttributeError:
            descr = ''
        return descr
    
    def get_skills(self):
        try:
            skills = [item.text for item in self.soup.find('div', class_='grid-block bullets-block').find_all('div', class_='bullets-block__item')]
        except:
            skills_html = self.soup.find_all('div', class_='portfolio-section__project')
            skills = []
            for skill in skills_html:
                title = skill.find('div', class_='portfolio-section__project-title').text
                descr = skill.find('div', class_='portfolio-section__project-description').text
                skills.append(':'.join((title, descr)))

        return ' | '.join(skills)
    

    def get_image(self):
        try:
            img = re.findall(".*?\(([^)]*)\).*", str(self.soup.find('div', class_='head-section__img')['style']))[0]
            return img
        except:
            pass
    def get_program(self):
        try:
            program_html = self.soup.find_all('div', class_='learning-program-section__item-content')
            if len(program_html) != 0:
                option = Options()
                # option.add_argument("--headless") # ФОНОВЫЙ РЕЖИМ

                browser = webdriver.Chrome(options=option)
                browser.get(self.course_url)
                browser.implicitly_wait(30)

                
                btn_see_more = browser.find_element(By.XPATH, '//div[@class="learning-program-section__item-circle-container learning-program-section__item-circle-container_hoverable"]')
                btn_see_more.click()
                time.sleep(2)

                selenium_soup = BeautifulSoup(browser.page_source, 'lxml')
                program_html = selenium_soup.find_all('div', class_='learning-program-section__item-content')
                lessons = []
                for item in program_html:
                    title = item.find('h3', class_='learning-program-section__item-title').text.replace(u'\xa0', u' ')
                    description = item.find('div', class_='learning-program-section__item-description').text.replace(u'\xa0', u' ')
                    lessons.append({
                        'name': title,
                        'descr': description.replace('\n\n\n', ' | ').replace('\n\n', '')
                    })
                result = {
                    'name':'',
                    'lessons': lessons
                }
                self.is_module = 'нет'
                return result    
            else:
                raise AttributeError

        except AttributeError:
            modules_html = self.soup.find_all('div', class_='curriculum-module curriculum-section__module')
            if len(modules_html) > 0:
                result = []
                for module in modules_html:
                    module_name = module.find('div', class_='curriculum-module__name').text
                    module_descr = module.find('div', class_='curriculum-module__description').text
                    module_content = module.find_all('div', class_='project-card')
                    module_content_tags = module.find_all('span', class_='tag curriculum-module__tag') # Некоторые темы указаны в виде тегов. Пример:https://practicum.yandex.ru/data-analyst/
                    module_data = []
                    if len(module_content) > 1:
                        for theme in module_content:
                            theme_name = theme.find('div', class_='project-card__name').text
                            theme_descr = re.sub('<div.*?\>', '', str(theme.find('div', class_='project-card__description').find('div', class_='paragraph'))).replace('</div>', '').replace('<br/>', ' | ')
                            module_data.append({
                                'name':theme_name,
                                'descr': theme_descr
                            })
                    elif len(module_content_tags) >= 1:
                        for tag in module_content_tags:
                            title = tag.text
                            module_data.append({
                                'name': title,
                                'descr': ''
                            })

                    result.append({
                        'name': module_name,
                        'descr': module_descr,
                        'lessons': module_data
                    })
                self.is_module = 'да'
                return result
            else:
                program_html = self.soup.find_all('section', class_='landing-grid demo-section__item')
                if len(program_html) > 0:
                    lessons = []
                    for item in program_html:
                        title = item.find(class_='demo-section__item-title').text
                        descr = item.find(class_='demo-section__item-description').text.replace('•', '').replace(u'\xa0', u'').split('\n')
                        filtered_descr = [i for i in descr if i != '']
                        lessons.append({
                            'name': title,
                            'descr': " | ".join(filtered_descr)
                        })
                    result = {
                        'name': '',
                        'lessons': lessons 
                    }
                    self.is_module = 'нет'
                    return result
                else:
                    browser = webdriver.Chrome()
                    browser.get(self.course_url)
                    browser.implicitly_wait(10)
                    time.sleep(20)

                    themes_buttons = browser.find_elements(By.XPATH, "//section[@id='zhizn-takomu-ne-uchila']//div[@class='lc-text-block lc-text-block_size_s16 lc-text-block_typeface_regular lc-tabs__tab-text']")
                    if len(themes_buttons) > 0:
                        lessons = [{'name':item.text, 'descr':''} for item in themes_buttons]
                        result = {
                            'name': '',
                            'lessons': lessons
                        }
                        self.is_module = 'нет'
                        return result
                    else:
                        print('Пустая программа', self.course_url)
                        self.is_module = 'нет'
                        return ''

date = datetime.now()
bot = YandexPractium(f'Practicum2({date.day}.{date.month}.{date.year}).xlsx')
bot.start()
# print(bot.get_program())