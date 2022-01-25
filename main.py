
import gspread
from oauth2client.service_account import ServiceAccountCredentials

import pandas as pd
import openpyxl  
from openpyxl import load_workbook
import xlrd
import time

from sys import platform

if platform == 'win32':
    import ctypes
    kernel32 = ctypes.windll.kernel32
    kernel32.SetConsoleMode(kernel32.GetStdHandle(-11), 7)



success_message = '\033[2;30;42m [SUCCESS] \033[0;0m' 
warning_message = '\033[2;30;43m [WARNING] \033[0;0m'
error_message = '\033[2;30;41m [ ERROR ] \033[0;0m'

class FinBot:
    """
    Класс ведет работу одновременно с гугл и иксель таблицами.\n
    Предоставляет три возможных действия работы:\n
        \t1. Расчет акции\n
        \t2. Стандартные цены\n
        \t3. Вход в акцию\n
    """
    def __init__(self):

        self.work_with_legal_entity()

        # НА ЗАМОРОЗКЕ 
        # method_load_data = input('Приветствуем! Давайте определимся откуда мы будем брать данные для расчёта:\n\n 1. Выберем какое-то конкретное юридическое лицо из списка\n 2. Загрузим собственный excel-файл с расчётами\n 0. Выход\n\n>>> ')
        
        # if method_load_data == '1':
        #     self.work_with_legal_entity()
        # elif method_load_data == '2':
        #     self.work_with_load_excel()
        # elif method_load_data == '0':
        #     quit()
        # else:
        #     print(error_message + '\tЯ Вас не понимаю. Введите мне в качестве ответа цифру 1 или 2')



    def work_with_legal_entity(self): # Работа с юр лицом
        # url = input('Укажите адрес вашего Google Sheet и предоставьте доступ для редактирования таблицы аккканту бота: helper@morbot-338716.iam.gserviceaccount.com')

        google_spread_url = '1bGbNieNgqDNSORaphLhLOHUbIUE00yxA0q_b4HsNclM' # ORIGINAL
        # google_spread_url = '1J6EJ601kR_S1_sMDFk4ibzMaG5mEgWUMcV1e_jL67Qs' # TEST
        spread = self.auth_spread(url=google_spread_url)

        list_legal_entity = spread.worksheets()[:5]
        legal_entity_with_pos_num = {}
        print(success_message+ '\tВот список доступных организаций. Выберите интересующую вас:')

        for num_pos, entity in enumerate(list_legal_entity):
            legal_entity_with_pos_num[num_pos+1] = entity.title
            print(f"{num_pos+1}. {entity.title}")
        print('0. Выход')

        choice_entity = input('>>> ')
        if choice_entity == '0': quit()

        for item in legal_entity_with_pos_num:
            if choice_entity == str(item):
                # print(spread.worksheet(legal_entity_with_pos_num[item]))
                print(success_message, f'\tВаш выбор: {legal_entity_with_pos_num[item]}\n')
                worksheet_url = spread.worksheet(legal_entity_with_pos_num[item])
                self.select_action(sheet_url=worksheet_url, table_url=google_spread_url)
    

    def select_action(self, sheet_url='0', table_url='0'):
        choice_action = input('Теперь выберем действие:\n 1. Расчёт акции (запись в googleSheet)\n 2. Вход в акцию\n 3. Стандартные цены (запись в excel)\n 0. Выход\n\n>>> ')

        if choice_action == '1':
            path_to_excel = input('Вставьте путь до файла "Отчёт по скидкам": ')
            self.set_caclulation_of_share(discount_report_file=path_to_excel, google_sheet=sheet_url, table_url=table_url)
            # print('Действие пока приостановлено')
        elif choice_action == '2':
            path_to_excel = input('Вставьте путь до файла "Отчёт по скидкам": ')
            self.entry_to_promotion(excel_file=path_to_excel, google_sheet=sheet_url)

        elif choice_action == '3':
            path_to_excel = input('Вставьте путь до файла "Отчёт по скидкам":\n>>> ')
            self.set_standart_price(discount_report_file=path_to_excel, google_sheet=sheet_url)
        elif choice_action == '0':
            quit()
        else:
            print(error_message + '\tЯ Вас не понимаю. Введите мне в качестве ответа цифру 1-3')


    def entry_to_promotion(self, excel_file, google_sheet=''):
        worksheet = google_sheet
        yes_or_no_col = worksheet.find('да/нет').col

        orders_ids = self.get_ids_in_excel(excel_file)
        
        def add_changes_in_excel(item):
            try:
                order_row = worksheet.find(str(item)).row
                order_status = worksheet.cell(order_row, yes_or_no_col).value
                if order_status == 'нет':
                    for row in worksheet_writer.iter_rows():
                        for j in row:
                            if str(j.value) == str(item):
                                # print('***********************************************')
                                print(warning_message + '\t Товар со статусом "Нет" (будет удален) ', item)
                                worksheet_writer.delete_rows(j.row)
                    
            
            except gspread.exceptions.APIError:
                print(error_message + '\tПревышен лимит запросов. Бот автоматически продолжит через 15 секунд ожидания.')
                time.sleep(20)
                add_changes_in_excel(item)

            except AttributeError:
                print(item)
        # excel_file = '/home/saloman/Documents/Business Automatization/Wildberries(Work)/BOT_finance/Отчет по скидкам для акции(не попадают под условие акции) (75) (2).xlsx'
        workbook_writer = openpyxl.load_workbook(excel_file)
        worksheet_writer = workbook_writer.worksheets[0]
        for item in orders_ids:
            add_changes_in_excel(item)
        
        workbook_writer.save(excel_file)

        print(success_message + f'\tТеперь в файле товары, только с актуальной информацией об акциях: {excel_file}')
        
        
    def get_ids_in_excel(self, excel_file):
        xlrd_workbook = xlrd.open_workbook(excel_file)
        xlrd_worksheet = xlrd_workbook.sheet_by_index(0)

        for item in range(xlrd_worksheet.nrows):
            try:
                row_title = xlrd_worksheet.col_values(item)[0] # Выбираем из списка конткретной колонки только первый элемент, он и будет у нас заглавием
                if row_title == 'Номенклатура (код 1С)':
                    excel_order_values = xlrd_worksheet.col_values(item)[1:] # Исключаем из списка колонки ячейку Номенклатура
                    excel_order_ids = [int(elem) for elem in excel_order_values] # Преобразовываем float --> int 
            
            except IndexError: # В конце итерации без использования этого исключения, вылетает ошибка IndexError
                break
        
        return excel_order_ids


    def set_caclulation_of_share(self, discount_report_file, google_sheet='0', table_url='0'): # Расчет акциии 
        def add_changes_to_sheet(item):
            try:
                order = workSheet_google.find(str(excel_ids[item])).row
                workSheet_google.update_cell(order, col_planned_price, str(excel_planned_price[item]))
                workSheet_google.update_cell(order, col_share_percentage, str(excel_percent_order[item]))
                print(success_message + '\tЗаписан ', item+1)
            except gspread.exceptions.APIError:
                print(error_message + '\tПревышен лимит запросов. Бот автоматически продолжит через 15 секунд ожидания.')
                time.sleep(20)
                add_changes_to_sheet(item)

        workBook_xrld = xlrd.open_workbook(discount_report_file)
        workSheet_xrld = workBook_xrld.sheet_by_index(0)

        for item in range(workSheet_xrld.ncols):
            row_title = workSheet_xrld.col_values(item)[0]
            if row_title == 'Номенклатура (код 1С)':
                excel_ids = [int(elem) for elem in workSheet_xrld.col_values(item)[1:]]
            elif row_title == 'Плановая цена для акции':
                excel_planned_price = [int(elem) for elem in workSheet_xrld.col_values(item)[1:]]
            elif row_title == 'Загружаемая скидка для участия в акции':
                excel_percent_order = [int(elem) for elem in workSheet_xrld.col_values(item)[1:]]
        
        workSheet_google = google_sheet
        
        col_planned_price = workSheet_google.find('Плановая цена').col
        col_share_percentage = workSheet_google.find('Требуемый %').col

        for item in range(len(excel_ids)):
            add_changes_to_sheet(item)

        print(success_message, f'\tАкции успешно расчитаны и записаны в Google Sheet по адресу: https://docs.google.com/spreadsheets/d/{table_url}')


    def set_standart_price(self, discount_report_file, google_sheet): # Раздел бота со стандартными ценами
        print(warning_message + '\tЗапустили запись стандартных цен в файл')
        def edit_price_and_share_in_excel_file(item):
            """
            Функция записи изменений в иксель из гугл таблиц
            """
            
            for item_with_row_num in range(workSheet_xlrd.nrows):

                try:
                    row_title = workSheet_xlrd.col_values(item_with_row_num)[0] # Выбираем из списка конткретной колонки только первый элемент, он и будет у нас заглавием
                    if row_title == 'Номенклатура (код 1С)':
                        ids = [int(i) for i in workSheet_xlrd.col_values(item_with_row_num)[1:]] # Повторное присвоение списка артикулов
                        for num_row in range(len(excel_order_ids)):
                            if str(excel_order_ids[num_row]) == str(excel_order_ids[item]): # Проверка для выявления номера строки артикула
                                item_row = num_row + 2 

                                order_row_in_googlesheet = workSheet_google.find(str(excel_order_ids[item])).row # Ищем артикул из иксель в гугл таблице, если его нет, сработает AttributeError
                                order_share = workSheet_google.cell(order_row_in_googlesheet, col_share_s_plus_a).value # Обращаемся к ячейке со строкой=строка товара и столбцом=столбец скидки
                                order_price = workSheet_google.cell(order_row_in_googlesheet, col_order_price).value # Обращаемся к ячейке со строкой=строка товара и столбцом=столбец цены
                                
                                order_price = order_price.split(',')[0]

                                worksheet_writer.cell(item_row, excel_new_price_before_share_col).value = order_price # Записываем в файл, но еще не сохраняем изменения
                                worksheet_writer.cell(item_row, excel_agreed_share_col).value = order_share # Записываем в файл, но еще не сохраняем изменения
                                
                                print(success_message, f'\t Записан {excel_order_ids[item]}')
                        break # Прекращаем поиск после нахождения колонки Номенклатура

                except gspread.exceptions.APIError:
                    print(error_message + '\tПревышен лимит запросов. Бот автоматически продолжит через 15 секунд ожидания.')
                    time.sleep(15)
                    edit_price_and_share_in_excel_file(item)

                except IndexError: # В конце итерации без использования этого исключения, вылетает ошибка IndexError
                    pass

                except AttributeError: 
                    print(warning_message, f'\tЭлемента нет в таблице расчётов {excel_order_ids[item]}')
        

        # Основной блок функции set_standart_price

        excel_filename = discount_report_file # Отчет по скидкам
        
        workSheet_google = google_sheet # Ссылка на страницу таблицы гугл

        col_order_price = workSheet_google.find('Цена прод').col # Берем номер стобца с информацией для цены товара
        col_share_s_plus_a = workSheet_google.find('Скидка(S+A)').col # Берем номер стобца с информацией для скидки товара

        # Модуль xlrd используем для поиска конкретного адреса значения
        # Благодаря модулю, мы можем найти номер колонки и номер строки товара
        workBook_xlrd = xlrd.open_workbook(excel_filename) 
        workSheet_xlrd = workBook_xlrd.sheet_by_index(0)

        # Модуль openpyxl используется для записи изменений в иксель
        # Плюсом модуля является то, что здесь можно указать ячейку для записи по номеру столбца и строки
        workbook_writer = openpyxl.load_workbook(excel_filename)
        worksheet_writer = workbook_writer.worksheets[0]

        # Поиск номеров колонок для записи
        for item in range(workSheet_xlrd.nrows):
            try:
                row_title = workSheet_xlrd.col_values(item)[0] # Выбираем из списка конткретной колонки только первый элемент, он и будет у нас заглавием
                if row_title == 'Номенклатура (код 1С)':
                    excel_order_values = workSheet_xlrd.col_values(item)[1:] # Исключаем из списка колонки ячейку Номенклатура
                    excel_order_ids = [int(elem) for elem in excel_order_values] # Преобразовываем float --> int 

                elif row_title == 'Новая розн. цена (до скидки)': # Условие нахождения колонки для записи в нее значений
                    excel_new_price_before_share_col = item + 1 # Если не прибавлять 1, то запись будет не попадать по нужной колонке. Будет записывать на колонку с номером item-1
                
                elif row_title == 'Согласованная скидка, %':
                    excel_agreed_share_col = item  + 1 # Если не прибавлять 1, то запись будет не попадать по нужной колонке. Будет записывать на колонку с номером item-1
            
            except IndexError: # В конце итерации без использования этого исключения, вылетает ошибка IndexError
                break

        # Поиск нужной ячейки по номеру колонки и строки и последующая за этим запись в ячейку
        for item in range(len(excel_order_ids)): # Пробегаемся по списку артикулов в таблице отчета по скидке 
            edit_price_and_share_in_excel_file(item)
            
        workbook_writer.save(excel_filename) # Сохраняем изменения
        print(success_message + '\tФайл изменен и сохранен ', excel_filename)
        

    def work_with_load_excel(self):
        pass


    def auth_spread(self, url):
        scope = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
        credentials = ServiceAccountCredentials.from_json_keyfile_name(r'Service Accounts/morbot-338716-b219142d9c70.json')

        gc = gspread.authorize(credentials)
        spread = gc.open_by_key(url)

        return spread

    def open_excel(self, input_name):
        """
        input_name Format: filename.xlsx
        output_name Format: filename without permission
        """

        excel = pd.read_excel(input_name)
        output_name = f'{input_name.split(".")[0]}.csv'
        excel.to_csv(output_name)
        file = pd.read_csv(output_name)

        return file


bot = FinBot()
