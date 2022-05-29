import json
import openpyxl
import requests
import xlrd
from datetime import datetime

from base import create_table


class Legends:
    def __init__(self, output_filename):
        self.result_filename = output_filename
        self.img_address = 'https://legend-lessons.s3.eu-central-1.amazonaws.com/'
        self.row_count = 2
        
        self.required_numbers_courses = 0
        self.run()

    def run(self):
        create_table(self.result_filename)
        count = 1
        while self.required_numbers_courses != 39:
            print('Пробуем проверить курс под номером ', count)
            self.url = f'https://urokilegend.ru/api/courses/{count}'
            data = self.get_course_json(self.url)
            if data != 500:
                print('Прошел проверку ', count)
                self.course_info = data['data']['course']
                self.add_row_to_table()
            count += 1
        print('ВСЁ СОБРАЛИ')

    def get_course_json(self, url):
        req = requests.get(url)
        status = req.status_code
        if status == 500:
            return status
        elif status == 200:
            json_data = json.loads(req.text)
            return json_data

    def get_price_json(self):
        req = requests.get(self.url+"/prices")
        prices_json = json.loads(req.text)
        return prices_json
    
    def add_row_to_table(self):
        workbook_writer = openpyxl.load_workbook(f'{self.result_filename}')
        sheet_writer = workbook_writer.worksheets[0]

        sheet_writer.cell(self.row_count, 1).value = 'Legends Lessons'
        sheet_writer.cell(self.row_count, 2).value = self.get_title()
        sheet_writer.cell(self.row_count, 5).value = self.get_description()
        sheet_writer.cell(self.row_count, 30).value = self.get_cover()
        sheet_writer.cell(self.row_count, 34).value = self.get_price()[0]
        sheet_writer.cell(self.row_count, 33).value = self.get_price()[1]
        sheet_writer.cell(self.row_count, 36).value = self.get_price()[2]
        sheet_writer.cell(self.row_count, 35).value = self.get_price()[3]
        sheet_writer.cell(self.row_count, 37).value = self.get_what_will_you_get()
        sheet_writer.cell(self.row_count, 40).value = self.get_program()
        sheet_writer.cell(self.row_count, 42).value = self.get_teachers()[0]
        sheet_writer.cell(self.row_count, 43).value = self.get_teachers()[1]
        sheet_writer.cell(self.row_count, 41).value = self.get_teachers()[2]
        sheet_writer.cell(self.row_count, 18).value = self.url.replace('/api', '') # !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        sheet_writer.cell(self.row_count, 20).value = self.get_start_date()

        workbook_writer.save(f'{self.result_filename}')

        self.row_count += 1 
        self.required_numbers_courses += 1

    def get_title(self):
        title = self.course_info['title']
        return title
    
    def get_description(self):
        description_course = " " if self.course_info['metaDescription'] == None else self.course_info['metaDescription']
        return description_course

    def get_cover(self):
        cover = self.img_address + self.course_info['courseInfoImage']
        return cover

    def get_start_date(self):
        start_date = " " if self.course_info['startDate'] == None else self.course_info['startDate'].split('T')[0] # форматируем такую строку 2022-03-08T00:00:00.000Z
        return start_date

    def get_what_will_you_get(self):
        what_will_you_get = [item['text'] for item in self.course_info['topics']]
        return " | ".join(what_will_you_get)

    def get_program(self):
        program_course = []
        for item in self.course_info['videos']:
            program_title = "<h3>" + item['title'] + "</h3>"
            program_description = " " if item['description'] == None else "<p>" + item['description'] + '</p>'
            program_course.append(program_title + program_description)
        return ''.join(program_course)
    
    def get_audience(self):
        audience = [f"<h3>{item['title']}</h3> <p>{item['description']}</p>" for item in self.course_info['audiences']]
        return " ".join(audience)

    def get_teachers(self):
        if self.course_info['speakerList'] == []:
                teacher_name = self.course_info['speaker']['fullName']
                teacher_description = self.course_info['description'] # совпадает описание курса и преподавателя
                teacher_avatar = self.img_address + self.course_info['speaker']['avatar']        

                return teacher_name, teacher_description, teacher_avatar
        else:
            list_teacher_name = [item['fullName'] for item in self.course_info['speakerList']]
            list_teacher_description = [item['description'] for item in self.course_info['speakerList']]
            list_teacher_avatar = [self.img_address+item['image'] for item in self.course_info['speakerList']]

            return tuple(map(lambda x: ' | '.join(x), (list_teacher_name, list_teacher_description, list_teacher_avatar)))
    
    def get_price(self):
        price_data = self.get_price_json()['data'][0]
        new_price = price_data['discount']['newPrice']
        old_price = price_data['discount']['oldPrice']
        installment_price = " " if price_data['installmentPrice'] == None else price_data['installmentPrice']
        installment_count = " " if price_data['installmentsCount'] == None else price_data['installmentsCount']

        return (new_price, old_price, installment_price, installment_count)


date = datetime.now()
bot = Legends(f'Legends({date.day}.{date.month}.{date.year}).xlsx')
# bot.run()
# print(bot.get_price())