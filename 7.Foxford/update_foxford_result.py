import xlrd
import openpyxl
from bs4 import BeautifulSoup

path = '/home/saloman/Documents/Edwica(Work)/Courses/7. FoxFord/Foxford(19.5.2022).xlsx'

def get_soup():
    file = open('foxford_source.html', 'r')
    soup = BeautifulSoup(file, 'lxml')
    file.close()
    return soup

def unlock_diapozone(text):
    text_parts = text.split('–')
    res = [i for i in range(int(text_parts[0]), int(text_parts[1].split()[0])+1)]
    return ', '.join(str(i) for i in res) + ' классы'

def change_levels():
    workbook_writer = openpyxl.load_workbook(path)
    worksheet_writer = workbook_writer.worksheets[0]

    wb_reader = xlrd.open_workbook(path)
    sh_reader = wb_reader.sheet_by_index(0)

    for row_num in range(1, sh_reader.nrows):
        old_level = sh_reader.cell(row_num, 10).value
        if '–' in old_level:
            new_level = unlock_diapozone(old_level)
            worksheet_writer.cell(row_num+1, 11).value = new_level
            print('Изменили ', new_level)
    workbook_writer.save(path)

def change_subcategories():
    workbook_writer = openpyxl.load_workbook(path)
    worksheet_writer = workbook_writer.worksheets[0]

    wb_reader = xlrd.open_workbook(path)
    sh_reader = wb_reader.sheet_by_index(0)

    soup = get_soup()
    all_courses = soup.find_all('div', class_='styled__Root-dkxgAg dSEYZS styled__CourseCard-buiJLf kNLeaB')
    
    for row_num in range(1, sh_reader.nrows):
        comparable_url = sh_reader.cell(row_num, 17).value
        # print(comparable_url)
        for course in all_courses:
            subcategory = course.find_all('div', class_='Text_root__3j40U Text_weight-normal__2T_yh Text_lineHeight-m__2YyYN Text_fontStyle-normal__264_c styled__Label-iPIFyI iKEUou FontSize_fontSize-m__3Cy9B Color_color-mineShaft__2PSyK')[-1].text
            course_url = "https://foxford.ru" + course.parent['href']
            if course_url == comparable_url:
                print('Записали на строке ', row_num+1)
                worksheet_writer.cell(row_num+1, 8).value = subcategory
    workbook_writer.save(path)

if __name__ == "__main__":
    change_levels()
    change_subcategories()
