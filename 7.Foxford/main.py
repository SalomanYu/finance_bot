from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import ElementClickInterceptedException

import json
import requests
import xlsxwriter
import openpyxl
from datetime import datetime

date = datetime.now()
filename = f'Foxford({date.day}.{date.month}.{date.year}).xlsx'

def get_soup():
    file = open('foxford_source.html', 'r')
    soup = BeautifulSoup(file, 'lxml')
    file.close()

    return soup

def create_table():
    workbook = xlsxwriter.Workbook(filename)
    sheet = workbook.add_worksheet()

    sheet.write(0, 0, 'Организатор курса')
    sheet.write(0, 1, 'Название курса')
    sheet.write(0, 2, 'Популярные')
    sheet.write(0, 3, 'Рекомендуемые')
    sheet.write(0, 4, 'Описание курса')
    sheet.write(0, 5, 'Выбор категории')
    sheet.write(0, 6, 'Язык программирования')
    sheet.write(0, 7, 'Выбор подкатегории')
    sheet.write(0, 8, 'Класс')
    sheet.write(0, 9, 'Предмет')
    sheet.write(0, 10, 'Уровень')
    sheet.write(0, 11, 'Баллы')
    sheet.write(0, 12, 'Сертификат')
    sheet.write(0, 13, 'Формат')
    sheet.write(0, 14, 'Тип')
    sheet.write(0, 15, 'Язык')
    sheet.write(0, 16, 'Продолжительность в неделях')
    sheet.write(0, 17, 'Ссылка на страницу')
    sheet.write(0, 18, 'Вариант для парсинга')
    sheet.write(0, 19, 'Дата начала')
    sheet.write(0, 20, 'Удостоверение о повышении клалификации')
    sheet.write(0, 21, 'Сертификат о прохождении курса')
    sheet.write(0, 22, 'Диплом об окончании')
    sheet.write(0, 23, 'Практико-ориентированность')
    sheet.write(0, 24, 'Наставник')
    sheet.write(0, 25, 'Помощь в трудоустройстве')
    sheet.write(0, 26, 'Наработка портфолио')
    sheet.write(0, 27, 'Рассрочка')
    sheet.write(0, 28, 'Налоговый вычет')
    sheet.write(0, 29, 'Обложка курса')
    sheet.write(0, 30, 'Картинка курса')
    sheet.write(0, 31, 'Приветственное видео')
    sheet.write(0, 32, 'Старая цена')
    sheet.write(0, 33, 'Цена')
    sheet.write(0, 34, 'Срок рассрочки в месяцах')
    sheet.write(0, 35, 'Платеж по рассрочке в рублях')
    sheet.write(0, 36, 'Навыки')
    sheet.write(0, 37, 'Программное обеспечение')
    sheet.write(0, 38, 'Модульный курс')
    sheet.write(0, 39, 'Программа курса')
    sheet.write(0, 40, 'Фото преподавателя')
    sheet.write(0, 41, 'ФИО преподавателя')
    sheet.write(0, 42, 'О преподавателе')
    sheet.write(0, 43, 'Описание преподавателя')
    sheet.write(0, 44, 'Для кого курс')
    sheet.write(0, 45, 'Преимущества')
    sheet.write(0, 46, 'Партнеры')
    sheet.write(0, 47, 'Лого партнеров')
    sheet.write(0, 48, 'Требуемые знания')

    workbook.close()


def add_row_to_table(data, row_num):
    workbook_writer = openpyxl.load_workbook(filename)
    worksheet_writer = workbook_writer.worksheets[0]

    # count = 3
    # for course in data:
    # try:
    worksheet_writer.cell(row_num, 1).value = 'Фоксфорд'
    worksheet_writer.cell(row_num, 2).value = data['title']
    worksheet_writer.cell(row_num, 5).value = data['description']
    worksheet_writer.cell(row_num, 6).value = 'Школа'
    worksheet_writer.cell(row_num, 8).value = data['subcategory']
    worksheet_writer.cell(row_num, 10).value = data['course_object']
    worksheet_writer.cell(row_num, 11).value = data['level']
    worksheet_writer.cell(row_num, 17).value = int(data['duration']) * 4
    worksheet_writer.cell(row_num, 18).value = data['url']
    worksheet_writer.cell(row_num, 31).value = data['course_img']
    worksheet_writer.cell(row_num, 32).value = data['video_url']
    worksheet_writer.cell(row_num, 34).value = data['price']['price_full']
    worksheet_writer.cell(row_num, 40).value = json.dumps(data['program'], ensure_ascii=False)
    # worksheet_writer.cell(row_num, 32).value = data['program']
    worksheet_writer.cell(
        row_num, 41).value = data['teachers']['teacher_avatar']
    worksheet_writer.cell(row_num, 42).value = data['teachers']['teacher_name']
    worksheet_writer.cell(
        row_num, 43).value = data['teachers']['teacher_description']
    # except KeyError:
    #     pass
    print(f"[{row_num-2}] Готов {data['title']}")
    workbook_writer.save(filename)


def parse_descriptions(url):
    file = open('desc.html', 'r')
    soup = BeautifulSoup(file, 'lxml')

    descriptions = soup.find_all('li', class_='styled__Item-cVjjrC IFfBX')
    result_data = []
    for desc in descriptions:
        try:
            p_tag = f'{desc.find("div", class_="Text_root__3j40U Text_weight-normal__2T_yh Text_lineHeight-l__2v_gr Text_fontStyle-normal__264_c styled__Description-ckRrQz jmOqCz FontSize_fontSize-xxxl__KbCq7 FontSize_fontSize-l-l__2Y-DV FontSize_fontSize-m-l__2cCSU FontSize_fontSize-s-l__10_XH Color_color-mineShaft__2PSyK").text}<br/>'
        except BaseException:
            p_tag = ''
        ul_tag = []
        for item in desc.find_all('li'):
            ul_tag.append(f'{item.text}<br/>')
        result_data.append(f'{p_tag} {" ".join(ul_tag)}')
    return result_data


def save_desc_page(page_source, driver):
    page_source = driver.page_source
    file = open('desc.html', 'w')
    file.write(page_source)
    file.close()


def get_course_program(url):
    options = Options()

    # options.add_argument("--headless") # ФОНОВЫЙ РЕЖИМ
    driver = webdriver.Chrome(options=options)
    driver.get(url.replace('/api', ''))
    driver.implicitly_wait(10)
    course_program_titles = [item.text for item in driver.find_elements(
        By.XPATH, '//div[@class="Text_root__3j40U Text_weight-normal__2T_yh Text_lineHeight-l__2v_gr Text_fontStyle-normal__264_c styled__Title-eWEYJA dgpnKP FontSize_fontSize-28__C5IlD FontSize_fontSize-l-28__1ho_8 FontSize_fontSize-m-24__cSDz7 FontSize_fontSize-xs-20__rruhO Color_color-inherit__27FNT"]')]


    try:
        for item in driver.find_elements(By.XPATH, "//div[@class='styled__Toggler-jiJmVh gawdid']"):
            item.click()
    except ElementClickInterceptedException:
        pass

    save_desc_page(driver.page_source, driver)
    driver.close()
    driver.quit()

    course_program_descriptions = parse_descriptions(url.replace('/api', ''))
    data = []
    for i in range(len(course_program_descriptions)):
        print(i)
        data.append({
            'name': course_program_titles[i],
            'desc': course_program_descriptions[i]
            })
    
    return data

def parse_course(url, row_num):
    headers = {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.99 Safari/537.36',
               "Content-Type": "application/json; charset=utf-8"}
    req = requests.get(url, headers=headers)
    try:
        data = json.loads(req.text.replace("'", '"'))
    except json.decoder.JSONDecodeError:
        print(url, ' JSON ERORR')
        return

    program = get_course_program(url)

    try:
        result_data = {
            'title': data['full_name'],
            'description': data['description'],
            'url': url.replace('/api', ''),
            'level': data['grade'],
            'course_object': data['name'],
            'subcategory': data['subtitle'],
            'duration': data['course_length'],
            'video_url': data['video_explainer_url'],
            'course_img': data['meta_image'],
            'price': {
                # 'price_in_month': data['user_cart_item']['price_per_month'],
                'price_full': data['user_cart_item']['price_with_discount']
                # except:

            },
            'teachers': {
                'teacher_name': f"{data['teachers'][0]['last_name']} {data['teachers'][0]['first_name']} { data['teachers'][0]['middle_name']}",
                'teacher_description': data['teachers'][0]['description'],
                'teacher_avatar': data['teachers'][0]['image']['large']
            },
            'program': {
                'name': '',
                'lessons': program
            }
        }
        add_row_to_table(result_data, row_num)
        print(result_data['title'])

    except KeyError:
        print(url)
        return

    # with open('result.json', 'w') as file:
    #     json.dump(result_data, file, ensure_ascii=False, indent=2)


def main():
    create_table()
    soup = get_soup()
    all_courses = soup.find_all(
        'div', class_='styled__Root-dkxgAg dSEYZS styled__CourseCard-buiJLf kNLeaB')

    result_data = []
    count = 2 
    for course in range(len(all_courses)):

        more_info_url = "https://foxford.ru/api" + \
            all_courses[course].parent['href']
        # result_data.append(parse_course(more_info_url, row_num=count))
        parse_course(more_info_url, row_num=count)
        print(f'[{count}] Записан курс')
        count += 1

    # add_row_to_table(result_data)


if __name__ == "__main__":
    main()
