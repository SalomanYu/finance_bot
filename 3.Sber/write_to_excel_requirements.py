import json
import xlrd
import xlsxwriter
import openpyxl

requirements = json.load(open('NO_DUBLICATE_REQ.json', 'r'))
# vacancies_with_req = json.load(open('group_vacancies.json', 'r'))

def set_up_matches(excel): # Установи соответствия

    writer_workbook = openpyxl.load_workbook(excel)
    writer_sheet = writer_workbook.worksheets[0]

    vacancies_groups  = json.load(open('groups_vacancies.json', 'r'))
    requirements = json.load(open('NO_DUBLICATE_REQ.json', 'r'))
    
    workbook = xlrd.open_workbook(excel)
    sheet = workbook.sheet_by_index(0)
    for group in vacancies_groups: # пригодится для сравнения по id для соответствия
        for item in range(1, sheet.nrows-1): # пробегаемся по всем строкам
            if group['id'] == int(sheet.cell(item, 0).value): # если нашли строку с таким id, как в группе, то берем у этой группы требования 
                for req_group in requirements:
                    if req_group['Группа'] == group['name']:
                        req = req_group['Требования'] # Взяли требования

                        for title_col in range(sheet.ncols): # пробегаемся по всем колонкам 
                            if sheet.cell(0, title_col).value in req: # если значение в колонке есть в требованиях группы, то записываемся в эту колонку и строку 1
                                writer_sheet.cell(item+1, title_col+1).value = 1
    writer_workbook.save('requirements_excel.xlsx')



def create_frame_table():
    vacancies = json.load(open('bigger_in_level2.json', 'r'))

    workbook = xlsxwriter.Workbook('excel_requirements.xlsx')
    worksheet = workbook.add_worksheet()

    worksheet.write(0, 0, 'Соответствия')
    worksheet.write(0, 1, 'Вес профессии в соответствии')
    worksheet.write(0, 2, 'Уровень должности')
    worksheet.write(0, 3, 'Вес профессии в уровне')
    worksheet.write(0, 4, 'Наименование професии и различные написания')

    col = 5
    for group in requirements:
        for req in group['Требования']:
            worksheet.write(0, col, req)
            col += 1

    row_count = 1
    for group in vacancies:
        for vacance in group['Группа']:
            try:
                weight_in_level = vacance['Вес в уровне']
                worksheet.write(row_count, 0, group['id'])
                worksheet.write(row_count, 1, vacance['Вес'])
                worksheet.write(row_count, 2, vacance['Уровень'])
                worksheet.write(row_count, 3, weight_in_level)
                worksheet.write(row_count, 4, vacance['Вакансия'])
                row_count += 1
            except KeyError:
                pass
    
    # set_up_matches('excel_requirements.xlsx')


    workbook.close()




def add_zeros(excel):
    workbook = xlrd.open_workbook(excel)
    sheet = workbook.sheet_by_index(0)

    writer_workbook = openpyxl.load_workbook(excel)
    writer_sheet = writer_workbook.worksheets[0]

    for num_col in range(5, sheet.ncols-5):
        for num_row in range(sheet.nrows):
            if writer_sheet.cell(num_row+1, num_col+1).value == None:
                writer_sheet.cell(num_row+1, num_col+1).value = 0
                print('добавили ноль')
            else:
                print(writer_sheet.cell(num_row+1, num_col+1).value)
    
    writer_workbook.save(excel)

                
create_frame_table()
# set_up_matches('excel_requirements.xlsx')
add_zeros('requirements_excel.xlsx')
