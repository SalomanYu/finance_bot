from bs4 import BeautifulSoup
import requests
import json
import re
import openpyxl
from datetime import datetime
from time import sleep

from base import create_table


class OnlineTutor:
    def __init__(self, outputfilename):
        self.outputfilename = outputfilename
        self.domain = 'https://www.tutoronline.ru/'
        self.row_count = 2

    def start(self):
        create_table(self.outputfilename)
        print('Успешно! Таблица создана')
        self.workbook_writer = openpyxl.load_workbook(self.outputfilename)
        self.sheet_writer = self.workbook_writer.worksheets[0]
        try:
            req = requests.get('https://www.tutoronline.ru/online-kursy')
            soup_page = BeautifulSoup(req.text, 'lxml')
        except requests.exceptions.ConnectionError:
            print('Ошибка интернет-соединения. Попробуем подключиться еще раз через минуту')
            sleep(60)

        all_courses = soup_page.find_all('a', class_='courses-list-card')
        for course in all_courses:
            self.title = course.find('div', class_='courses-list-card__title').text.strip()
            
            print(f"{[self.row_count-1]} {self.title}")

            self.url = self.domain + course['href']
            self.teacher_avatar = course.find('div', class_='courses-list-card__teacher-avatar').img['src']
            self.teacher_name = course.find('div', class_='courses-list-card__teacher-name').text.strip()            
            try:
                self.course_object, self.level = course.find('div', class_='courses-list-card__subject').text.strip().split(',')
            except ValueError:
                self.course_object = course.find('div', class_='courses-list-card__subject').text.strip()
                self.level = ''
            
            if self.course_object == 'Пакетное предложение':
                self.course_object = ''
                self.parse_complex_course(self.url)
                self.add_complex_to_table()
                continue
            try:
                request = requests.get(self.url)
                self.soup = BeautifulSoup(request.text, 'lxml')
                self.add_row_to_table()
            except requests.exceptions.ConnectionError:
                print('Ошибка интернет-соединения. Попробуем подключиться еще раз через минуту')
                sleep(60)
            # break
        self.workbook_writer.save(self.outputfilename)
        print('Успешно! Таблица сохранена')

    def add_row_to_table(self):
        self.sheet_writer.cell(self.row_count, 1).value = 'Online Tutor'
        self.sheet_writer.cell(self.row_count, 2).value = self.title
        self.sheet_writer.cell(self.row_count, 5).value = self.__get_description().replace('✔', '')
        self.sheet_writer.cell(self.row_count, 6).value = 'Школа'
        self.sheet_writer.cell(self.row_count, 11).value = self.level
        self.sheet_writer.cell(self.row_count, 16).value = 'Русский'
        self.sheet_writer.cell(self.row_count, 18).value = self.url
        self.sheet_writer.cell(self.row_count, 40).value = json.dumps(self.get_program(self.url), ensure_ascii=False)

        self.sheet_writer.cell(self.row_count, 33).value = self.__get_price()[0]
        self.sheet_writer.cell(self.row_count, 34).value = self.__get_price()[1]
        self.sheet_writer.cell(self.row_count, 41).value = self.teacher_avatar
        self.sheet_writer.cell(self.row_count, 42).value = self.teacher_name
        self.sheet_writer.cell(self.row_count, 43).value = self.__get_about_teacher()
        self.sheet_writer.cell(self.row_count, 46).value = self.__get_advantages()

        self.row_count += 1   

    def add_complex_to_table(self):

        description, about_teacher_filtered, old_price, price, advantages_result, modules = self.parse_complex_course(self.url)

        self.sheet_writer.cell(self.row_count, 1).value = 'Online Tutor'
        self.sheet_writer.cell(self.row_count, 2).value = self.title
        self.sheet_writer.cell(self.row_count, 5).value = description.replace('✔', '')
        self.sheet_writer.cell(self.row_count, 6).value = 'Школа'
        self.sheet_writer.cell(self.row_count, 11).value = self.level
        self.sheet_writer.cell(self.row_count, 16).value = 'Русский'
        self.sheet_writer.cell(self.row_count, 18).value = self.url
        # self.sheet_writer.cell(self.row_count, 20).value = self.startDate
        self.sheet_writer.cell(self.row_count, 40).value = json.dumps(modules, ensure_ascii=False)


        # Teacher info
        # name, description, image = self.get_teacher_info()

        # self.sheet_writer.cell(self.row_count, 31).value = self.course_image
        self.sheet_writer.cell(self.row_count, 33).value = old_price
        self.sheet_writer.cell(self.row_count, 34).value = price
        self.sheet_writer.cell(self.row_count, 41).value = self.teacher_avatar
        self.sheet_writer.cell(self.row_count, 42).value = self.teacher_name
        self.sheet_writer.cell(self.row_count, 43).value = about_teacher_filtered
        self.sheet_writer.cell(self.row_count, 46).value = advantages_result
        # self.sheet_writer.cell(self.row_count, 37).value = self.get_skills()

        self.row_count += 1         

    def parse_complex_course(self, url):
        # url = 'https://www.tutoronline.ru/kursy-paketnoe-predlozhenie/paket-kursov-podgotovki-dlya-10-klassa'
        req = requests.get(url)
        complex_soup = BeautifulSoup(req.text, 'lxml')

        # Описание курса
        try:
            description = complex_soup.find('div', class_='b-course-page-header-info__time-row').text.strip().replace('\n', '<br/>')
        except:
            description = ''
        # Об Учителе
        try:
            about_teacher = complex_soup.find('div', class_='b-course-page-header-teacher__info-details').text.strip().split('\n\n')
            about_teacher_filtered = '<br/>'.join(about_teacher).replace('\n', ' ')
        except:
            about_teacher_filtered = ''
        # Цена
        try:
            prices = complex_soup.find_all('div', class_='b-course-page-subscribe-price-row b-course-page-subscribe-price-row__primary')[-1].text.split()
            old_price = prices[0]
            price = prices[1]
        except:
            price, old_price = ''

        # Преимущества
        try:
            all_advantages = complex_soup.find('div', class_='b-course-page-text').find_all('li')
            advantages_result = []
            for item in all_advantages:
                header = item.b.text
                content = item.span.text
                advantages_result.append(f"{header}:{content}")
        except:
            advantages_result = []

        try:
            courses_included_in_complex = complex_soup.find_all('div', class_='l-row-shell b-course-page-title-and-text-wrap')
            modules = []
            for item in courses_included_in_complex:
                name = item.h2.text.strip()
                url = self.domain + item.a['href']
                program = self.get_program(url)
                program['name'] = name
                modules.append(program)
        except:
            modules = ''
        return description, about_teacher_filtered, old_price, price, ' | '.join(advantages_result), modules
        

    def __get_description(self):
        try:
            description = self.soup.find('div', class_='b-course-page-header-info__time-row').text.strip().replace('\n', '<br/>')
        except:
            description = ''
        return description
    
    def __get_price(self):
        try:
            prices = self.soup.find_all('div', class_='b-course-page-subscribe-price-row b-course-page-subscribe-price-row__primary')[-1].text.split()
            old_price = prices[0]
            price = prices[1]
            return old_price, price
        except IndexError:
            return '', ''
    def __get_about_teacher(self):
        try:
            advantages_resultall_info = self.soup.find('div', class_='b-course-page-header-teacher__info-details').text.strip().split('\n\n')
        except:
            advantages_resultall_info = ''
        return '<br/>'.join(advantages_resultall_info).replace('\n', ' ')

    def __get_advantages(self): # Преимущества
        try:
            all_advantages = self.soup.find_all('div', class_='courses-list')[-1].find_all('div', class_='courses-list-card')
            results = []
            for item in all_advantages:
                advantage = item.text.strip().replace('\n', ':')
                results.append(advantage)

            return " | ".join(results)
        except:
            return ''
    
    def get_program(self, url):
        req = requests.get(url)
        soup = BeautifulSoup(req.text, 'lxml')
        program_html = soup.find_all('details', class_='b-course-page-accordion-item')
        lessons = []
        for item in program_html:
            header_with_whitespaces = item.summary.text.strip()
            header = re.sub(' +', ' ', header_with_whitespaces).replace('\n', ' ')
            try:
                content = item.ul.text.strip().split('\n')
            except AttributeError:
                content = ''
            lessons.append({
                'name': header.replace(u'\r', u''),
                'desc': " | ".join(content)
            })
        program = {
            'name': '',
            'lessons': lessons
        }
        return program
date = datetime.now()
bot = OnlineTutor(f'OnlineTutor({date.day}.{date.month}.{date.year}).xlsx')
bot.start()

