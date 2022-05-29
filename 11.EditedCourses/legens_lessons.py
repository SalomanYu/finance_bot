import re
import json

import openpyxl

import numpy as np
import pandas as pd

from base import create_table

class LEGEND:
    def __init__(self, excelfilename, outputfilename):
        self.input_excel = excelfilename
        self.outputfilename = outputfilename
        self.row_count = 2

        create_table(self.outputfilename)
        self.connect_to_table()

    def run(self):
        for course_index in range(len(self.all_descriptions)):
            self.course_index = course_index
            self.add_course_to_result_table()
        print('ВСЁ')
    
    def add_course_to_result_table(self):
        workbook_writer = openpyxl.load_workbook(self.outputfilename)
        sheet_writer = workbook_writer.worksheets[0]

        sheet_writer.cell(self.row_count, 1).value = 'Уроки Легенд'
        sheet_writer.cell(self.row_count, 11).value = 'Начальный'
        sheet_writer.cell(self.row_count, 16).value = 'Русский'
        sheet_writer.cell(self.row_count, 2).value = self.all_names[self.course_index]
        sheet_writer.cell(self.row_count, 5).value = self.all_descriptions[self.course_index] 
        sheet_writer.cell(self.row_count, 31).value= self.all_directions[self.course_index]
        sheet_writer.cell(self.row_count, 33).value = self.all_old_price[self.course_index]
        sheet_writer.cell(self.row_count, 34).value = self.all_course_prices[self.course_index]
        sheet_writer.cell(self.row_count, 35).value = self.all_rassrochka_month[self.course_index]
        sheet_writer.cell(self.row_count, 36).value = self.all_rassrochka_price[self.course_index]
        sheet_writer.cell(self.row_count, 48).value = self.all_min_khowledges[self.course_index] 
        sheet_writer.cell(self.row_count, 37).value = self.all_skills[self.course_index] 
        
        sheet_writer.cell(self.row_count, 40).value = json.dumps(self.edit_course_program_form(self.all_programs[self.course_index]), ensure_ascii=False)

        sheet_writer.cell(self.row_count, 41).value = self.all_teacher_avatars[self.course_index]
        sheet_writer.cell(self.row_count, 42).value = self.all_teacher_names[self.course_index]
        sheet_writer.cell(self.row_count, 43).value = self.all_teacher_about[self.course_index]
        sheet_writer.cell(self.row_count, 44).value = self.all_teacher_descriptions[self.course_index]
        
        sheet_writer.cell(self.row_count, 18).value = self.all_urls[self.course_index]
        sheet_writer.cell(self.row_count, 31).value = self.all_course_images[self.course_index]

        sheet_writer.cell(self.row_count, 20).value = self.format_date(str(self.all_startDates[self.course_index]))
        

        self.row_count += 1
        workbook_writer.save(self.outputfilename)
        print(self.all_names[self.course_index])



    def connect_to_table(self):
        workbook = pd.ExcelFile(self.input_excel)
        sheet = workbook.parse(0)

        self.all_names = list(sheet['Название курса'])
        self.all_directions = list(sheet['Продолжительность в неделях'])
        self.all_urls = list(sheet['Ссылка на страницу'])
        self.all_startDates = list(sheet['Дата начала'])
        self.all_course_images = list(sheet['Картинка курса'])
        self.all_old_price = list(sheet['Старая цена'])
        self.all_course_prices = list(sheet['Цена'])
        self.all_rassrochka_month = list(sheet['Срок рассрочки в месяцах'])
        self.all_rassrochka_price = list(sheet['Платеж по рассрочке в рублях'])
        self.all_teacher_avatars = list(sheet['Фото преподавателя'])
        self.all_teacher_names = list(sheet['ФИО преподавателя'])
        self.all_teacher_names = list(sheet['ФИО преподавателя'])
        self.all_teacher_about = list(sheet['О преподавателе'])
        self.all_teacher_descriptions = list(sheet['Описание преподавателя'])
        
        self.all_descriptions = list(sheet['Описание курса'])
        self.all_skills = list(sheet['Навыки'])
        self.all_programs = list(sheet['Программа курса'])
        self.all_min_khowledges = list(sheet['Требуемые знания'])
    
    def format_date(self, date):
            if date == 'NaT':
                return ''
            else:
                date_list = date.split(' ')[0].split('-')
                reverse_date = '.'.join(date_list[::-1])
                return reverse_date
    
    
    def edit_course_program_form(self, data):
        pattern_headers = '<h3>.*?\</h3>'
        pattern_contents = "<p>.*?\</p>|<p>.*?\\n</p>"
        all_modules_titles = re.findall(pattern_headers, data)
        all_modules_content = re.split(pattern_headers, data)[1:]

        module_data = []
        for content in range(len(all_modules_content)):
            themes = re.findall(pattern_contents, all_modules_content[content])
            
            module_title = re.sub('<.*?\>', '', all_modules_titles[content])
            # module_content = []
            module_content = {}
            for theme in themes:
                module_content['name'] = module_title
                module_content['desc'] = re.sub('<.*?\>', '', theme).replace('\n', '')
                
            module_data.append(module_content)   
        result = {
            'name': '',
            'lessons': module_data
        }
        return result

bot = LEGEND("Non-edited files/Legends(21.5.2022).xlsx", "Results/Legends(21.5.2022).xlsx")

bot.run()