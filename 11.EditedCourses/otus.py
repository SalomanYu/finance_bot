from termios import ECHOE
import openpyxl
import pandas as pd
from base import create_table
import re
import json

class OTUS:
    def __init__(self, excelfilename, outputfilename):
        self.input_excel = excelfilename
        self.outputfilename = outputfilename
        self.row_count = 2

        create_table(self.outputfilename)
        self.connect_to_table()

    def run(self):
        # print(self.format_skills(self.clear_tags(self.all_min_khowledges[59])))
        for course_index in range(len(self.all_descriptions)):
            self.course_index = course_index
            self.add_course_to_result_table()
            # break
        
        print('ВСЁ')
    
    def add_course_to_result_table(self):
        workbook_writer = openpyxl.load_workbook(self.outputfilename)
        sheet_writer = workbook_writer.worksheets[0]

        sheet_writer.cell(self.row_count, 1).value = 'OTUS'
        sheet_writer.cell(self.row_count, 11).value = 'Продвинутый'
        sheet_writer.cell(self.row_count, 6).value = 'IT'
        sheet_writer.cell(self.row_count, 13).value = 'Да'
        sheet_writer.cell(self.row_count, 14).value = 'Онлайн'
        sheet_writer.cell(self.row_count, 15).value = 'Курс'
        sheet_writer.cell(self.row_count, 16).value = 'Русский'
        sheet_writer.cell(self.row_count, 24).value = 'Да'
        sheet_writer.cell(self.row_count, 25).value = 'Да'
        sheet_writer.cell(self.row_count, 27).value = 'Да'
        sheet_writer.cell(self.row_count, 29).value = 'Да'
        sheet_writer.cell(self.row_count, 2).value = self.all_names[self.course_index]
        sheet_writer.cell(self.row_count, 5).value = self.clear_tags(self.all_descriptions[self.course_index]) 
        sheet_writer.cell(self.row_count, 30).value = self.format_image_links(self.all_covers[self.course_index])
        sheet_writer.cell(self.row_count, 31).value= self.all_directions[self.course_index]
        sheet_writer.cell(self.row_count, 34).value = self.all_course_prices[self.course_index]
        sheet_writer.cell(self.row_count, 49).value = self.format_skills(self.clear_tags(self.all_min_khowledges[self.course_index]))

        sheet_writer.cell(self.row_count, 37).value = self.format_skills(self.clear_tags(self.all_skills[self.course_index])) 
        sheet_writer.cell(self.row_count, 40).value = json.dumps(self.edit_course_program_form(self.all_programs[self.course_index]), ensure_ascii=False)

        sheet_writer.cell(self.row_count, 41).value = self.format_image_links(self.all_teacher_avatars[self.course_index])
        sheet_writer.cell(self.row_count, 42).value = self.all_teacher_names[self.course_index]
        sheet_writer.cell(self.row_count, 43).value = self.all_teacher_about[self.course_index]
        sheet_writer.cell(self.row_count, 44).value = self.all_teacher_descriptions[self.course_index]
        
        sheet_writer.cell(self.row_count, 48).value = self.format_image_links(self.all_partners[self.course_index])
        sheet_writer.cell(self.row_count, 18).value = self.all_urls[self.course_index]
        sheet_writer.cell(self.row_count, 31).value = self.format_image_links(self.all_course_images[self.course_index])
        sheet_writer.cell(self.row_count, 20).value = self.change_date_format(self.all_startDates[self.course_index])


        self.row_count += 1
        workbook_writer.save(self.outputfilename)

    def connect_to_table(self):
        workbook = pd.ExcelFile(self.input_excel)
        sheet = workbook.parse(0)

        self.all_names = list(sheet['Название курса'])
        self.all_directions = list(sheet['Продолжительность в неделях'])
        self.all_urls = list(sheet['Ссылка на страницу'])
        self.all_startDates = list(sheet['Дата начала'])
        self.all_covers = list(sheet['Обложка курса'])
        self.all_course_images = list(sheet['Картинка курса'])
        self.all_course_prices = list(sheet['Цена'])
        self.all_teacher_avatars = list(sheet['Фото преподавателя'])
        self.all_teacher_names = list(sheet['ФИО преподавателя'])
        self.all_teacher_names = list(sheet['ФИО преподавателя'])
        self.all_teacher_about = list(sheet['О преподавателе'])
        self.all_teacher_descriptions = list(sheet['Описание преподавателя'])
        self.all_partners = list(sheet['Партнеры'])
        
        self.all_descriptions = list(sheet['Описание курса'])
        self.all_skills = list(sheet['Навыки'])
        self.all_programs = list(sheet['Программа курса'])
        self.all_min_khowledges = list(sheet['Требуемые знания'])

    def change_date_format(self, date):
        date_dict = {
            '01':('январь', 'января', 'январе'),
            '02': ('февраль', 'февраля', 'феврале'),
            '03': ('март', 'марта', 'марте'),
            '04': ('апрель', 'апреля', 'апреле'),
            '05': ('май', 'мая', 'мае'),
            '06': ('июнь', 'июня', 'июне'),
            '07': ('июль', 'июля', 'июле'),
            '08': ('август', 'августа', 'августе'),
            '09': ('сентябрь', 'сентября', 'сентябре'),
            '10': ('октябрь', 'октября', 'октябре'),
            '11': ('ноябрь', 'ноября', 'ноябре'),
            '12': ('декабрь', 'декабря', 'декабре'),
        }
        try:
            day, month = date.split(' ')
            for key, value in date_dict.items():
                if month in value:
                    if day.isdigit():
                        return f'{day}.{key}.2022'
                    else:
                        return f"01.{key}.2022"
        except:
            return ''

    def format_image_links(self, data):
        try:
            return data.replace('(', '').replace(')', '').replace("'", '')
        except:
            return ''

    def clear_tags(self, data):
        without_span = re.sub("<span.*?\>", ' ', str(data)).replace('</span>', ' ')
        without_headers = re.sub("</h\d>", "</p><br/>", re.sub('<h\d>', '<p>', without_span))
        without_lists_class = re.sub('<li.*?\>', '', without_headers)
        without_lists = without_lists_class.replace('<li>', '').replace('</li>', '<br/>').replace('<ul>', '').replace('</ul>', '')
        result = re.sub('<strong>|</strong>|<b>|</b>', '', without_lists)
        if result == 'nan':
            print(data)
        return result
    
    def format_skills(self, data):
        stop_words = ('<p>Обязательно:</p><br/> ', '<p>Будет плюсом:</p><br/>', '<br/>Будет плюсом:<br/><br/>',
                    '<p>Этот курс вам подойдёт, если вы: <br/>', '</p>', '<p>Для обучения вам понадобится базовый опыт программирования на Python, а именно, следующие знания: <br/><br/>',
                    '<p>Для успешного обучения и оптимального усвоения уроков вы должны знать:<br/><br/>',
                    '<p>Требования к поступающим </p>', ' Обязательно: ', 'Желательно: <br/>',
                    '<p>Вам будет гораздо проще учиться, если вы: <br/>', '<p>Вам будет гораздо проще учиться, если вы: <br/>',
                    '<p>Для прохождения программы необходимы:<br/>', 'Необходимое:', 'Программирование:<br/>',
                    'Технологии:<br/><br/>', '<p>Курс рассчитан на:</p>')
        # cleared_skills = data
        for item in stop_words:
            if item in data:
                data=data.replace(item, '')
        
        skill_list = re.split('<br/>', data.replace('<p>', '').replace('</p>', '<br/>'))
        result = []

        for item in skill_list:
            if len(item) > 2:
                result.append(item)
        return ' | '.join(result)
        # print(result)

    def edit_course_program_form(self, data):
        # result = []
        # for item in range(len(self.all_programs)):
        pattern_headers = '<h3>.*?\</h3>'
        pattern_contents = "<p>.*?\</p>"
        all_modules_titles = re.findall(pattern_headers, data)
        all_modules_content = re.split(pattern_headers, data)[1:]
       
        module_data = []
        for content in range(len(all_modules_content)):
            themes = re.findall(pattern_contents, all_modules_content[content])
            
            module_title = re.sub('<.*?\>', '', all_modules_titles[content])
            module_content = []
            for theme in themes:
                module_content.append({
                    'name': re.sub('<.*?\>', '', theme), 
                    'desc': ''
                })
                
            module_data.append({
                'name': module_title, 
                'lessons': module_content
                })    
        return module_data
            # result.append({
                # item: module_data
                # })
            # break


bot = OTUS('Non-edited files/OTUS(21.5.2022).xlsx', 'Results/OTUS(21.5.2022).xlsx')
bot.run()