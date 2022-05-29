import pandas as pd
import re
import json
from base import create_table
import openpyxl

class Uchmet:
    def __init__(self, excelfilename, outputfilename):
        self.input_excel = excelfilename
        self.outputfilename = outputfilename
        self.row_count = 2

        create_table(self.outputfilename)
        self.connect_to_table()

    def run(self):
        for course_index in range(len(self.all_names)):
            self.course_index = course_index
            self.add_course_to_result_table()
            print(self.all_names[self.course_index])
            # break
        print('ВСЁ')
    
    def add_course_to_result_table(self):
        workbook_writer = openpyxl.load_workbook(self.outputfilename)
        sheet_writer = workbook_writer.worksheets[0]

        sheet_writer.cell(self.row_count, 40).value = json.dumps(self.edit_program(self.all_programs[self.course_index]),
                                                                 ensure_ascii=False) 
        sheet_writer.cell(self.row_count, 1).value = 'Uchmet'
        sheet_writer.cell(self.row_count, 2).value = self.all_names[self.course_index]
        sheet_writer.cell(self.row_count, 5).value = self.all_descriptions[self.course_index]
        sheet_writer.cell(self.row_count, 14).value = self.all_format[self.course_index]
        sheet_writer.cell(self.row_count, 15).value = self.all_types[self.course_index]
        sheet_writer.cell(self.row_count, 16).value = 'Русский'
        sheet_writer.cell(self.row_count, 18).value = self.all_urls[self.course_index]
        sheet_writer.cell(self.row_count, 22).value = 'Да'
        sheet_writer.cell(self.row_count, 31).value = self.all_course_images[self.course_index]
        sheet_writer.cell(self.row_count, 34).value = self.all_course_prices[self.course_index]

        sheet_writer.cell(self.row_count, 41).value = self.all_teacher_avatars[self.course_index]
        sheet_writer.cell(self.row_count, 42).value = self.all_teacher_names[self.course_index]
        sheet_writer.cell(self.row_count, 43).value = self.all_teacher_about[self.course_index]
        

        self.row_count += 1
        workbook_writer.save(self.outputfilename)

    def connect_to_table(self):
        workbook = pd.ExcelFile(self.input_excel)
        sheet = workbook.parse(0)

        self.all_names = list(sheet['Название курса'])    
        self.all_programs = list(sheet['Программа курса'])
        self.all_urls = list(sheet['Ссылка на страницу'])
        self.all_format = list(sheet['Формат'])        
        self.all_types = list(sheet['Тип'])
        self.all_course_images = list(sheet['Картинка курса'])
        self.all_course_prices = list(sheet['Цена'])
        self.all_teacher_avatars = list(sheet['Фото преподавателя'])
        self.all_teacher_names = list(sheet['ФИО преподавателя'])
        self.all_teacher_names = list(sheet['ФИО преподавателя'])
        self.all_teacher_about = list(sheet['О преподавателе'])        
        self.all_descriptions = list(sheet['Описание курса'])
    
    def edit_program(self, data):
        try:
            program_elems = re.findall(r'\d\d\. .*?\n|\d\. .*?\n', data) # Находим строки начинающиеся на 1. Название
            if  len(program_elems) > 0:
                program_data = []
                for item in program_elems:
                    edited_item = item.replace(item.split(' ')[0], '') + '<br/>'
                    program_data.append({
                        'name': edited_item.replace('\n', '').strip(),
                        'desc': ''
                    })
                result = {
                    'name': '',
                    'lessons': program_data
                }
                return result
            else:
                program_elems = data.replace('<li>', '').replace('</li>', '<br/>').replace('<h3>', '').replace('</h3>', '<br/>')
                data = []
                for item in program_elems.split('<br/>')[:-1]:
                    data.append({
                        'name': item.strip(),
                        'desc': ''
                    })
                result = {
                    'name': '',
                    'lessons': data
                }
                
                return result
        except:
            return ''

bot = Uchmet('Non-edited files/Uchmet(23.5.2022).xlsx', 'Results/Uchmet(23.5.2022).xlsx')
bot.run()