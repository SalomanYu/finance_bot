import json
import re
import openpyxl

from base import create_table

class SF:
    def __init__(self, outputfilename):
        self.outputfilename = outputfilename
        self.row_count = 2
    
    def start(self):
        create_table(self.outputfilename)
        print('Успешно! Таблица создана')
        self.workbook_writer = openpyxl.load_workbook(self.outputfilename)
        self.sheet_writer = self.workbook_writer.worksheets[0]

        all_courses = json.load(open('all_courses.json', 'r'))
        for course in all_courses:
            self.name = all_courses[course]['name']
            self.description = all_courses[course]['short_desc']
            self.price = all_courses[course]['price']
            # self.category = all_courses[course]['faculty']
            self.url = all_courses[course]['url']
            self.image = all_courses[course]['image_url']
            self.program = self.get_program(all_courses[course]['description'])

            self.add_row_to_table()
        self.workbook_writer.save(self.outputfilename)

    def get_program(self, data):
        direction_pattern = 'Длительность: .*?\\n'
        pattern = '\d{2}. .*?\\n|\d{1}. .*?\\n'
        without_direction = re.sub(direction_pattern, '', data)
        headers = re.findall(pattern, without_direction)
        descriptions = re.split(pattern, without_direction)[1:]

        data = []
        for item in range(len(descriptions)):
            data.append({
                'name': headers[item].replace('\n', ''),
                'desc': descriptions[item].strip().replace('\n\n\n', ' | ').replace('\n', ' | ')
            })

        result = {
            'name': '',
            'lessons': data
        }
        return result

    def add_row_to_table(self):
        self.sheet_writer.cell(self.row_count, 1).value = 'SF Education'
        self.sheet_writer.cell(self.row_count, 2).value = self.name
        self.sheet_writer.cell(self.row_count, 5).value = self.description
        self.sheet_writer.cell(self.row_count, 6).value = 'Финансы'
        self.sheet_writer.cell(self.row_count, 16).value = 'Русский'
        self.sheet_writer.cell(self.row_count, 18).value = self.url
        self.sheet_writer.cell(self.row_count, 40).value = json.dumps(self.program, ensure_ascii=False)
        
        self.sheet_writer.cell(self.row_count, 31).value = self.image
        self.sheet_writer.cell(self.row_count, 34).value = self.price

        self.row_count += 1


bot = SF('hello.xlsx')
bot.start()
