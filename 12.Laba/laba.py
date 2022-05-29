import json
import requests
from bs4 import BeautifulSoup, Tag
import re
import openpyxl
from datetime import datetime
from base import create_table


class Laba:
    def __init__(self, outputfilename):
        self.outputfilename = outputfilename
        self.row_count = 2

    def start(self):
        # req = requests.get('https://l-a-b-a.com/lecture/1921-praktika-biznes-analiza-v-it')
        # req = requests.get('https://l-a-b-a.com/lecture/743-google-tablicy')
        # req = requests.get('https://l-a-b-a.com/lecture/show/149')
        # req = requests.get('https://l-a-b-a.com/lecture/1558-upravlenie-tehnicheskimi-proektami')
        # self.soup = BeautifulSoup(req.text, 'lxml')
        # print(self.get_program())
        # print(self.get_title())
        # print(self.get_description())
        # print(self.get_audience())
        # print(self.get_level())
        # print(self.get_teacher_info())
        
        create_table(self.outputfilename)        
        for page_num in range(1, 9):
            url = f'https://l-a-b-a.com/lecture/{page_num}'
            req = requests.get(url)
            self.soup = BeautifulSoup(req.text, 'lxml')
            self.parse_course_list()
    
    def parse_course_list(self):
        page_courses = self.soup.find_all('div', class_='col-lg-4 col-md-6 col-12 courses__item')
        for course in page_courses:
            self.course_category = course.find('div', class_='label label__course').text.strip()
            self.course_url = course.a['href'] # 'https://l-a-b-a.com/' + 
            if 'https://l-a-b-a.com/' not in self.course_url:
                self.course_url = 'https://l-a-b-a.com' + self.course_url
            print(f"[{self.row_count-1}] {self.course_url}")
            
            if self.course_url == 'https://l-a-b-a.com/lecture/course-packages':
                print('Курсы закончились')
                quit()

            req = requests.get(self.course_url)
            self.soup = BeautifulSoup(req.text, 'lxml')
            self.add_to_table()
            

    def add_to_table(self):
        workbook_writer = openpyxl.load_workbook(self.outputfilename)
        sheet_writer = workbook_writer.worksheets[0]

        sheet_writer.cell(self.row_count, 1).value = 'LABA'
        sheet_writer.cell(self.row_count, 2).value = self.get_title()
        sheet_writer.cell(self.row_count, 5).value = self.get_description()
        sheet_writer.cell(self.row_count, 8).value = self.course_category
        sheet_writer.cell(self.row_count, 11).value = self.get_level()
        sheet_writer.cell(self.row_count, 16).value = 'Русский'
        sheet_writer.cell(self.row_count, 18).value = self.course_url
        sheet_writer.cell(self.row_count, 40).value = json.dumps(self.get_program(), ensure_ascii=False)

        # Teacher info
        name, description, image = self.get_teacher_info()

        sheet_writer.cell(self.row_count, 31).value = image
        sheet_writer.cell(self.row_count, 41).value = image
        sheet_writer.cell(self.row_count, 42).value = name
        sheet_writer.cell(self.row_count, 43).value = description
        sheet_writer.cell(self.row_count, 45).value = self.get_audience()

        self.row_count += 1
        workbook_writer.save(self.outputfilename)
        

    def get_title(self):
        try:
            title = self.soup.find('div', class_='promo__title seo__item wow fadeInLeft').text.strip()
        except AttributeError:
            try:
                title = self.soup.find('div', class_='l_intro-title').h1.text.strip()
            except:
                title = self.soup.find('h1', class_='hero__title').text.strip()
        
        return re.sub(' +', ' ', title).replace('\n', ' ')

    def get_description(self):
        result = []
        try:
            try:
                description_html = self.soup.find('section', class_='sc__topics sc__bg-1 landing-redesign').find('div', class_='details__list')
            except AttributeError:
                try:
                    description_html = self.soup.find('section', class_='sc__topics sc__bg-1 landing-redesign updated').find('div', class_='details__list')
                except AttributeError:
                    try:
                        description_html = self.soup.find('section', class_='sc__topics sc__bg-1 landing-redesign min-topics').find('div', class_='details__list')
                    except AttributeError:
                        description_html = self.soup.find_all('li', class_='pros__list-item')
                        if len(description_html) == 0:
                            raise AttributeError     
                        for item in description_html:
                            if isinstance(item, Tag):
                                header = f"<p>{item.find('p', class_='pros__list-item__lead').text}</p>"
                                descr = f"<p>{item.find_all('p')[-1].text}</p>"
                                result.append(header+descr)
                        return ' '.join(result)

            for item in description_html:
                if isinstance(item, Tag):
                    header = f"<p>{item.find('div', class_='details__item-header').text}</p>"
                    descr = f"<p>{item.find('div', class_='details__item-descr').text}</p>"
                    result.append(header+descr)
        except AttributeError: 
            description_html = self.soup.find_all('div', class_='program-item')
            # print(description_html)
            for item in description_html:
                try:
                    header = f"<p>{item.find('h3', class_='program-title with-plus').text}</p>"
                except:
                    header = f"<p>{item.find('span', class_='program-title with-plus').text}</p>"

                descr = f"<p>{item.p.text}</p>"
                result.append(header+descr)  
        return ' '.join(result)
    
    def get_level(self):
        try:
            level = self.soup.find('div', class_='promo__marker-title').text
        except AttributeError:
            level = ''
        return level

    def get_audience(self):
        result = []
        try:
            audience_html = self.soup.find('div', class_='side__right-box side-box-recommend').find('div', class_='rec__list')
            for item in audience_html:
                if isinstance(item, Tag):
                    header_with_whitespaces = item.find('div', class_='rec__item-title wow fadeInLeft').text
                    header = re.sub(' +', ' ', header_with_whitespaces).replace('\n', ' ')
            
                    descr_with_whitespaces = item.find('div', class_='rec__item-descr wow fadeInLeft').text
                    descr = re.sub(' +', ' ', descr_with_whitespaces).replace('\n', ' ')
                    result.append(" : ".join((header, descr)))
            return " | ".join(result)
        except AttributeError:
            try:
                audience_html = self.soup.find('ul', class_='cells__grid')
                if audience_html == None:
                    raise AttributeError
                for item in audience_html:
                    if isinstance(item, Tag):
                        header = item.find('header', class_='cell__lead').text
                        descr_with_whitespaces = item.find('div', class_='cell__content reg-sm').text
                        descr = re.sub(' +', ' ', descr_with_whitespaces).replace('\n', ' ')

                        result.append(" : ".join((header, descr)))
                return " | ".join(result)
            
            except AttributeError:
                return ''
    

    def get_teacher_info(self):
        try:
            name = self.soup.find('div', class_='teacher__name wow fadeInLeft').text
            about_teacher = self.soup.find('div', class_='teacher__position wow fadeInLeft').text
            img = 'https://l-a-b-a.com/' + self.soup.find('div', class_='teacher__img wow fadeInRight').img['src']
        except AttributeError:
            try:
                try:
                    name = self.soup.find('h3', class_='l_teacher-name').text
                except AttributeError:
                    name = self.soup.find('h4', class_='l_teacher-name').text

                about_teacher = self.soup.find('div', class_='l_teacher-intro').p.text
                img = 'https://l-a-b-a.com/' + self.soup.find('div', class_='l_teacher-photo').img['data-src']
            except AttributeError:
                name = self.soup.find('p', class_='hero__lector caps-md').find_all('span')[0].text
                about_teacher = self.soup.find('p', class_='hero__lector caps-md').find_all('span')[-1].text
                img = 'https://l-a-b-a.com/' +self.soup.find('div', class_='hero__img-wrap').img['src']


        return name, about_teacher, img

    def get_program(self):
        course_content = []
        
        try:
            program_html = self.soup.find_all('div', class_='program__item wow fadeInRightFull')
            
            # print(program_html)
            if len(program_html) == 0:
                program_html = self.soup.find('div', class_='program__list wow fadeInRightFull')
                if len(program_html) == 0:
                    raise TypeError

            for item in program_html:
                if isinstance(item, Tag):
                    try:
                        label = item.find('div', class_='program__label').text
                        if 'Модуль' in label:
                            continue
                        header_with_whitespaces = item.find('div', class_='program__title').text
                        header = re.sub(' +', ' ', header_with_whitespaces).replace('\n', ' ')
                        try:
                            descr_with_whitespaces = item.find('div', class_='program__descr-text').text
                            descr = re.sub(' +', ' ', descr_with_whitespaces).replace('\n\n', '').replace('\n', '\n ')
                            if descr == '\n ':
                                descr = ''
                        except AttributeError:
                            descr = ''
                    except AttributeError: # ХЗ почему возникает тут ошибка, но ладно...
                        break

                    
                    course_content.append({
                        'name': str(label + header).replace(u'\ufeff', u''),
                        'desc': descr
                    })
            result = {
                'name': '',
                'lessons': course_content
            }
        
            return result

        except TypeError:

            try:
                program_html = self.soup.find_all('li', class_='l_course-main-item')
                
                if len(program_html) == 0:
                    raise AttributeError
                for item in program_html:
                    if isinstance(item, Tag):
                        label = item.find('p', class_='item-order').text
                        
                        try:
                            header_with_whitespaces = item.find('span', class_='item-title').text
                        except AttributeError:
                            header_with_whitespaces = item.find('h6', class_='item-title').text
                        
                        header = re.sub(' +', ' ', header_with_whitespaces).replace('\n', ' ')
                        
                        try:
                            descr = item.find('ul', class_='accordion-body accordion-wp-body').text
                            if descr == '\n ':
                                    descr = ''
                        except AttributeError:
                            descr = ''
                        
                        course_content.append({
                            'name': '.'.join((label,header)).replace(u'\ufeff', u''),
                            'desc': descr
                        })
            except:
                program_html = self.soup.find_all('li', class_='programm__list-item prog-item js-accordeon-item')
                for item in program_html:
                    if isinstance(item, Tag):
                        header = item.find('p', class_='prog-item__lead reg-lg js-accordeon-trigger').text
                        descr_with_whitespaces = item.find('div', class_='prog-item__body js-accordeon-body').text
                        descr = re.sub(' +', ' ', descr_with_whitespaces).replace('\n\n', '').replace('\n', ' | ')
                        course_content.append({
                            'name': header,
                            'desc': descr
                        })
    
            result = {
                'name': '',
                'lessons': course_content
            }
            return result
                    
date = datetime.now()
bot = Laba(f'LABA({date.day}.{date.month}.{date.year}).xlsx')
bot.start()
# bot.parse_course_list()