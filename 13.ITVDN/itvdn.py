from bs4 import BeautifulSoup
import requests
import json
import openpyxl
import re
from datetime import datetime

from base import create_table

class ITVDN:
    def __init__(self, outputfilename):
        self.outputfilename = outputfilename
        self.row_count = 2
        self.domain = 'https://itvdn.com/ru'

    def start(self):
        # req = requests.get('https://itvdn.com/ru/video/start')
        # self.soup = BeautifulSoup(req.text, 'lxml')
        create_table(self.outputfilename)
        all_courses = json.load(open('all_courses.json', 'r'))
        for item in range(len(all_courses)):
            self.course_name = all_courses[item]['Title']
            self.course_url = self.domain + all_courses[item]['CourseUrl']
            self.course_image = self.domain + all_courses[item]['ImageUrl']
            self.course_level = all_courses[item]['DifficultyLevel']
            self.course_descripton = all_courses[item]['Description']
            self.is_free = all_courses[item]['IsFree']
            self.startDate = all_courses[item]['Modified']
            if self.is_free:
                self.price = ''
            else:
                self.price = '2238'

            self.req = requests.get(self.course_url)
            self.soup = BeautifulSoup(self.req.text, 'lxml')
            print(f'[{self.row_count-1}] {self.course_name}: {self.course_url}')
            
            self.add_row_to_table()
            # break
    
    def add_row_to_table(self):
        workbook_writer = openpyxl.load_workbook(self.outputfilename)
        sheet_writer = workbook_writer.worksheets[0]

        sheet_writer.cell(self.row_count, 1).value = 'ITVDN'
        sheet_writer.cell(self.row_count, 2).value = self.course_name
        sheet_writer.cell(self.row_count, 5).value = self.course_descripton
        sheet_writer.cell(self.row_count, 6).value = 'IT'
        sheet_writer.cell(self.row_count, 11).value = self.course_level
        sheet_writer.cell(self.row_count, 16).value = 'Русский'
        sheet_writer.cell(self.row_count, 18).value = self.course_url
        sheet_writer.cell(self.row_count, 20).value = self.startDate
        sheet_writer.cell(self.row_count, 40).value = json.dumps(self.get_program(), ensure_ascii=False)


        # Teacher info
        name, description, image = self.get_teacher_info()

        sheet_writer.cell(self.row_count, 31).value = self.course_image
        sheet_writer.cell(self.row_count, 34).value = self.price
        sheet_writer.cell(self.row_count, 41).value = image
        sheet_writer.cell(self.row_count, 42).value = name
        sheet_writer.cell(self.row_count, 43).value = description
        sheet_writer.cell(self.row_count, 37).value = self.get_skills()

        self.row_count += 1
        workbook_writer.save(self.outputfilename)


    def get_title(self):
        try:
        
            title = self.soup.find('div', class_='about-course-header header-with-breadcrumbs').h1.text
            return title
        except:
            print(self.course_url)
            quit()
    def get_teacher_info(self):
        try:
            name = self.soup.find('a', class_='header-3 author-name-link').text
            descr = self.soup.find('p', itemprop='description').text
            img = self.soup.find('meta', itemprop='image')['content']
            return name, descr, img
        except:
            # try:
            name = self.soup.find('p', class_='author-name-link').text
            descr = self.soup.find('p', class_='description-author-courses').text
            img = self.soup.find('meta', itemprop='image')['content']
            return name, descr, img
            # except BaseException as err:
                # print(err)
                # print(self.course_url)
                # quit()


    def get_program(self):
        lessons = self.soup.find_all('div', class_='lsn-name-wrapper')[:-1]
        content = []
        for item in lessons:
            content.append({
                'name': item.text,
                'desc': ''
            })
        result = {
            'name': '',
            'lessons': content
        }
        return result
    
    def get_description(self):
        description_html = self.soup.find('div', class_='course-description scroll-style').find_all('p')[:-1]
        description = []
        stop_words = 'вы научитесь'
        for item in description_html:
            paragraph = item.text.replace(u'\xa0', u'')
            if stop_words in  paragraph.lower() :
                continue
            description.append("<p>" + paragraph + "</p>")
        return ''.join(description)
    
    def get_skills(self):
        try:
            skills_ul = [item.text for item in self.soup.find('div', class_='course-description scroll-style').find_all('li')]
        except:
            try:
                skills_ul = [item.text for item in self.soup.find('div', class_='learn-course symbols').find_all('li')]
            except:
                return ""
        return " | ".join(skills_ul)


date = datetime.now()
bot = ITVDN(f'ITVDN2({date.day}.{date.month}.{date.year}).xlsx')
bot.start()
# bot.get_welcome_video()
# print('Название:\n', bot.get_title(), end='\n_______________\n')
# print('Учитель:\n', bot.get_teacher_info(), end='\n_______________\n')
# print('Описание:\n', bot.get_description(), end='\n_______________\n')
# print('Программа:\n', bot.get_program(), end='\n_______________\n')
# print('Скиллы:\n', bot.get_skills(), end='\n_______________\n')