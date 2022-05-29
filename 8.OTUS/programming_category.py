from bs4 import BeautifulSoup
import requests
import re
import json
import openpyxl
from base import create_table

class Course:
    def __init__(self, result_filename):
        self.result_filename = result_filename
        self.row_count = 2 # В openpyxl отчет строк и стобцов с 1, а на первой строке у нас заголовки

        self.categories = (
            'https://otus.ru/categories/programming/', 'https://otus.ru/categories/operations/',
            'https://otus.ru/categories/data-science/', 'https://otus.ru/categories/gamedev/',
            'https://otus.ru/categories/marketing-business/', 'https://otus.ru/categories/testing/')

    def run(self):
        create_table(self.result_filename)

        for category_url in self.categories:
            req = requests.get(category_url)
            soup = BeautifulSoup(req.text, 'lxml')

            lessons = soup.find('div', class_='lessons__page').find_all('a')
            for course in lessons:      
                self.url = "https://otus.ru" + course['href']
                self.course_img = course.img['src']
                in_development = course.find('div', class_='lessons__new-item-discount lessons__new-item-discount_greenlight')

                if in_development == None: # Если НЕТ плажки 'В разработке', то продолжаем
                    try:
                        req_course = requests.get(self.url)
                        self.soup = BeautifulSoup(req_course.text, 'lxml')
        
                        self.add_course_to_result_table()
                    except AttributeError as err:
                        # pass
                        print(err, self.url)
                        # file = open(f'ERRORS/{self.result_filename}.txt', 'a')
                        # file.write(f"{self.url} \n")
                        # file.close()


    def add_course_to_result_table(self):
        workbook_writer = openpyxl.load_workbook(self.result_filename)
        sheet_writer = workbook_writer.worksheets[0]
        descr = self.get_description_course() + self.get_graduation_project() + self.get_learning_process()
        sheet_writer.cell(self.row_count, 1).value = 'OTUS'
        sheet_writer.cell(self.row_count, 2).value = self.get_course_title()
        sheet_writer.cell(self.row_count, 5).value =  descr
        sheet_writer.cell(self.row_count, 30).value = self.get_cover_course()
        sheet_writer.cell(self.row_count, 17).value= self.get_direction()
        sheet_writer.cell(self.row_count, 34).value = self.get_price().split('₽')[0]
        sheet_writer.cell(self.row_count, 49).value = self.get_min_knowledge()
        sheet_writer.cell(self.row_count, 37).value = self.get_what_will_you_get()
        sheet_writer.cell(self.row_count, 40).value = self.get_course_program()
        sheet_writer.cell(self.row_count, 42).value = ' | '.join(self.get_teachers()[0])
        sheet_writer.cell(self.row_count, 44).value = ' | '.join(self.get_teachers()[1])
        sheet_writer.cell(self.row_count, 41).value = ' | '.join(self.get_teachers()[2])
        sheet_writer.cell(self.row_count, 46).value = self.get_partners()
        sheet_writer.cell(self.row_count, 18).value = self.url
        sheet_writer.cell(self.row_count, 31).value = self.course_img
        sheet_writer.cell(self.row_count, 20).value = self.get_start_date()
        sheet_writer.cell(self.row_count, 43).value = ' | '.join(self.get_teachers()[3])

        print(self.row_count)
        self.row_count += 1
        workbook_writer.save(self.result_filename)


    def get_course_title(self):
        try:
            title_course = self.soup.find('div', class_='course-header2__title').text.strip()
        except AttributeError:
            try:
                title_course = self.soup.find('h1', class_='course-header2__title').text.strip()
            except AttributeError:
                title_course = self.soup.find('div', class_='preparatory-intro__title').text


        return title_course

    def get_description_course(self):
        try:
            # 2. ЗАГОЛОВОК ОПИСАНИЯ 
            title_description = "<h3>" + self.soup.find('div', class_='course-about__title').text.strip() + "</h3>"
            # 2.1 ОСНОВНОЕ ОПИСАНИЕ
            description = self.soup.find('div', class_='course-about__content')  # Поработать с тегами

            # <div class="course-about__content"> Убираем всё такое
            filtered_description = re.sub("<div.*?\>", '', str(description)).replace('</div>', '').replace('\n', '').replace('<li>', '<p>').replace('</li>', '</p>')
        except AttributeError:
            about_course = str(self.soup.find('div', class_='preparatory-intro__list')).replace('<div class="preparatory-intro__list">', "").replace('</div>', '').replace('<li>', '<p>').replace('</li>', '</p>').strip()
            return about_course

        return " ".join((title_description, filtered_description))

    
    def get_graduation_project(self):
        try:
            # 2.2 ВЫПУСКНОЙ ПРОЕКТ - ЗАГОЛОВОК
            graduation_project_title = '<h3>' + self.soup.find('div', class_='course-graduation-project__title').text + '</h3>'
            # 2.3 ВЫПУСКНОЙ ПРОЕКТ - СОДЕРЖАНИЕ
            graduation_project_content = '<p>' + self.soup.find('div', class_='course-graduation-project__content').text + '</p>'
        except AttributeError:
            graduation_project_title = graduation_project_content = ''
            print('\tОтсутствует Выпускной проект')
        
        return ' '.join((graduation_project_title, graduation_project_content))

    
    def get_learning_process(self):
        try:
            # 2.4 ПРОЦЕСС ОБУЧЕНИЯ - ЗАГОЛОВОК
            learning_process_title = '<h3>' + self.soup.find('div', class_='course-process-description__title').text.strip() + '</h3>'
            # 2.5 ПРОЦЕСС ОБУЧЕНИЯ - СОДЕРЖАНИЕ
            learning_process_content = str(self.soup.find('div', class_='course-process-description__content')).replace('<div class="course-process-description__content">', '').replace('</div>', '').replace('<li>', '<p>').replace('</li>', '</p>')
        except AttributeError:
            learning_process_content = learning_process_title = ''
            print('\tОтсутствует Процесс обучения')

        return ' '.join((learning_process_title, learning_process_content))

    def get_cover_course(self):
        try:
            cover_course = re.findall("([^']*)", self.soup.find('div', class_='course-header2__image-container')['style'])[2]
        except (IndexError, TypeError):
            cover_course = ''

        return cover_course
    

    def get_direction(self):
        try:
            # 4. ДЛИТЕЛЬНОСТЬ КУРСА
            direction_course_container = self.soup.find('div', class_='course-header2-bottom__content-item container__col container__col_4 container__col_md-4 container__col_ssm-12')
            direction_course = direction_course_container.find('p', class_='course-header2-bottom__item-text').text.strip()
            try:
                direction_course = int(direction_course.split('\n')[0]) * 4 # Делим строку по энтеру и берем первое число(количество месяцев) умножаем на количество недель
            except ValueError:
                direction_course = ''
        except AttributeError:
            direction_course = self.soup.find('div', class_='preparatory-payment-list__item-caption').text.strip()

        return direction_course


    def get_start_date(self):
        # 5. ДАТА НАЧАЛА
        try:
            # start_date = self.soup.find_all('p', class_='course-header2-bottom__item-text')[2].text.strip()
            start_date_container = self.soup.find('div', class_='course-header2-bottom__content-item container__col container__col_3 container__col_md-3 container__col_ssm-12')
            start_date = start_date_container.find('p', class_='course-header2-bottom__item-text').text.strip()
        except AttributeError:
            start_date = ''
            print('\tОтсутствует Дата начала')

        return start_date 


    def get_teachers(self):
        try:    
            # Прописываем сплит, чтобы убрать инфу о месте работы
            teacher_names = [self.soup.find('div', class_='course-teacher__name').text.strip()] + [item.text.strip().split('\n')[0] for item in self.soup.find_all('div', class_='course-teacher hide js-teacher')]
            # Соединяем описание основного препода с остальными
            teacher_descriptions = [self.soup.find('div', class_='js-teacher-dsc course-teacher__about').text.strip()] + [item.text.strip() for item in self.soup.find_all('div', class_='js-teacher-dsc course-teacher__about hide js-teacher')]
            teacher_avatars = [re.findall("\(.*?\)",item['style'])[0] for item in self.soup.find('div', class_='container__col container__col_3 container__col_sm-12').find_all('div', class_='course-teacher__photo ic ic-avatar-stub')]
            
            try:
                teacher_works = [self.soup.find('div', class_='course-teacher__work').text, ]
                hide_teachers = self.soup.find_all('div', class_='course-teacher hide js-teacher')
                for teacher in hide_teachers:
                    work = teacher.find('div', class_='course-teacher__work')
                    if work == None:
                        work = ''
                        teacher_works.append(work)
                        continue
                    teacher_works.append(work.text)
            except AttributeError:
                teacher_works = ['' for i in range(len(teacher_names))]


        except AttributeError:
            teacher_names = [item.p.text.strip() for item in self.soup.find_all('div', class_='hide-ssm preparatory-teachers-card__sign')]
            teacher_descriptions = [item.text.strip() for item in self.soup.find_all('p', class_='preparatory-teachers-card__text')]
            teacher_avatars = [re.findall("\(.*?\?",item['style'])[0] for item in self.soup.find_all('div', class_='preparatory-teachers-card__image hide-ssm ic')]
            teacher_works = ['' for i in range(len(teacher_names))]

        return teacher_names, teacher_descriptions, teacher_avatars, teacher_works

    def get_min_knowledge(self):
        # 7. НЕОБХОДИМЫЕ ЗНАНИЯ
        min_knowledge = str(self.soup.find('div', class_='course-min__content')).replace('<div class="course-min__content">', '').replace('</div>', '').replace('\n', '')
        if min_knowledge in ('None', None):
            min_knowledge = ''
        return min_knowledge

    def get_course_program(self):
        # 8. ПРОГРАММА КУРСА
        course_program_info = self.soup.find_all('div', class_='preparatory-program-list-item')
        filtered_course_program_info = []
        
        for module in course_program_info:
            themes = []
            
            title_module = "<h3>" + module.find('span', class_='preparatory-program-list-item__module-title').text + "</h3>" # Название модуля
            theme_count = ["<p> " + item.text for item in module.find_all('span', 'preparatory-program-list-item__theme')] # Блоки "Тема №"
            theme_name = ["<strong> " + item.text + " </strong> </p>" for item in module.find_all('span', 'preparatory-program-list-item__title')] # Блоки с названием темы
            
            for item in range(len(theme_count)):
                themes.append(theme_count[item] + theme_name[item]) # Приводим к виду "Тема 1. Название темы" 
            
            filtered_course_program_info.append(title_module + ' '.join(themes)) # Приводим к виду "Название модуля \n все темы этого модуля"  
        return ' '.join(filtered_course_program_info)


    def get_what_will_you_get(self):
        try:
            # 9. ЧТО ВЫ ПОЛУЧИТЕ ПОСЛЕ ОБУЧЕНИЯ - ЗАГОЛОВОК
            # title_what_will_you_get = '<h3>' + self.soup.find('div', class_='course-left__title').text.strip() + '</h3>'
            # 9.1. ЧТО ВЫ ПОЛУЧИТЕ ПОСЛЕ ОБУЧЕНИЯ - СОДЕРЖИМОЕ
            what_will_you_get = self.soup.find('div', class_='container__col container__col_7 container__col_sm-12 course-left__sm-padding').find('div', class_='course-left__content')
            filtered_what_will_you_get =  re.sub("<div.*?\>", "", str(what_will_you_get)).replace("</div>", "")
        except AttributeError:
            # title_what_will_you_get = filtered_what_will_you_get = ''
            filtered_what_will_you_get = ''
            print('\tОтсутствует После обучения вы')
        return filtered_what_will_you_get

    def get_partners(self):
        # 10. ПАРНЕРЫ - КАРТИНКИ
        partners = [re.findall("\(.*?\)", partner['style'])[0] for partner in self.soup.find_all('div', class_='course-partners__item-logo')]
        return ' | '.join(partners)


    def get_price(self):
        # 11. СТОИМОСТЬ
        try:
            price = self.soup.find('div', class_='course-bottom-bar-meta__value').text.strip().replace('\n', ' ')
        except AttributeError:
            try:
                price = self.soup.find('div', class_='preparatory-intro__main-price__text').text
            except AttributeError:
                price = ''
                print('\tОтсутствует Цена')
        return price


from datetime import datetime
date = datetime.now()
bot = Course(f'OTUS({date.day}.{date.month}.{date.year}).xlsx')
bot.run()


